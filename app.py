from flask import (
    Flask, render_template, request, send_file,
    jsonify, after_this_request, session, redirect, g,
    send_from_directory,
)
from flask_wtf.csrf import CSRFProtect, CSRFError
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv
import os, uuid, time, logging, threading, secrets, shutil
import requests as _requests
from pathlib import Path
from translation_service import translate_text as _ts_translate_text
from datetime import datetime as _datetime

# ── Security Manager (Quota, Auth, DDoS Protection) ─────────────────────────
from security_manager import (
    get_client_fingerprint,
    ConversionCounter,
    TranslationQuota,
    ProToken,
    VoucherSecurity,
    CaptchaVerifier,
    require_captcha,
    check_conversion_quota,
    LogSanitizer,
    SanitizingLogger,
)
import jwt  # For JWT token handling

# ── RTL (Arabic / Hebrew) rendering helpers ────────────────────────────────
try:
    import arabic_reshaper as _arabic_reshaper
    from bidi.algorithm import get_display as _bidi_get_display
    _RTL_LIBS_OK = True
except ImportError:
    _RTL_LIBS_OK = False

_RTL_LANG_NAMES = {"Arabic", "Hebrew"}

def _is_rtl_lang(lang_name: str) -> bool:
    return lang_name in _RTL_LANG_NAMES

def _prepare_rtl_text(text: str) -> str:
    """Reshape + reorder Arabic/Hebrew text for correct PDF glyph rendering."""
    if not _RTL_LIBS_OK:
        return text
    try:
        return _bidi_get_display(_arabic_reshaper.reshape(text))
    except Exception:
        return text

load_dotenv(override=True)

# ── Logging with Password Sanitization ────────────────────────────────────
log_level = logging.DEBUG if os.getenv("FLASK_DEBUG", "False").lower() == "true" else logging.INFO
logging.basicConfig(
    level=log_level,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("app.log", encoding="utf-8"),
    ],
)
# Use SanitizingLogger to automatically scrub passwords from logs
logger = SanitizingLogger(__name__)

# ── LibreOffice detection ───────────────────────────────────────────────────
LIBREOFFICE_PATH = shutil.which("libreoffice") or shutil.which("soffice")
if LIBREOFFICE_PATH:
    logger.info("LibreOffice found at: %s", LIBREOFFICE_PATH)
else:
    logger.warning("LibreOffice NOT found — Word to PDF will be unavailable")


# ── App setup ──────────────────────────────────────────────────────────────
app = Flask(__name__)
app.debug = os.getenv("FLASK_DEBUG", "False").lower() == "true"
app.config["UPLOAD_FOLDER"]      = os.getenv("UPLOAD_FOLDER", "uploads")
app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MAX_CONTENT_LENGTH", 32 * 1024 * 1024))
app.config["MAX_FILE_SIZE"]      = int(os.getenv("MAX_FILE_SIZE",      32 * 1024 * 1024))
app.config["FILE_EXPIRY_HOURS"]  = int(os.getenv("FILE_EXPIRY_HOURS",  24))
app.config["SECRET_KEY"]         = os.getenv("SECRET_KEY", "dev-key-change-in-production")
app.config["CONVERSION_TIMEOUT"] = int(os.getenv("CONVERSION_TIMEOUT", 120))

# ── Startup security check ─────────────────────────────────────────────────
_is_production = os.getenv("FLASK_ENV", "production").lower() == "production"
if _is_production and app.config["SECRET_KEY"] == "dev-key-change-in-production":
    raise RuntimeError(
        "SECRET_KEY must be set to a strong random value in production. "
        "Set the SECRET_KEY environment variable."
    )

session_cookie_secure = os.getenv("SESSION_COOKIE_SECURE")
if session_cookie_secure is None:
    app.config["SESSION_COOKIE_SECURE"] = os.getenv("FLASK_ENV", "production").lower() == "production"
else:
    app.config["SESSION_COOKIE_SECURE"] = session_cookie_secure.lower() == "true"
app.config["SESSION_COOKIE_HTTPONLY"] = os.getenv("SESSION_COOKIE_HTTPONLY", "True").lower() == "true"
app.config["SESSION_COOKIE_SAMESITE"] = os.getenv("SESSION_COOKIE_SAMESITE", "Lax")

# ── PayPal ─────────────────────────────────────────────────────────────────
PAYPAL_CLIENT_ID      = os.getenv("PAYPAL_CLIENT_ID", "")
PAYPAL_CLIENT_SECRET  = os.getenv("PAYPAL_CLIENT_SECRET", "")
PAYPAL_MODE           = os.getenv("PAYPAL_MODE", "sandbox")   # "sandbox" or "live"
PAYPAL_PRICE_USD      = os.getenv("PAYPAL_PRICE_USD", "2.00")
PAYPAL_API_BASE       = (
    "https://api-m.sandbox.paypal.com"
    if PAYPAL_MODE == "sandbox"
    else "https://api-m.paypal.com"
)
FREE_CONVERSIONS_LIMIT  = int(os.getenv("FREE_CONVERSIONS_LIMIT", 3))
PAID_CONVERSIONS_AMOUNT = int(os.getenv("PAID_CONVERSIONS_AMOUNT", 20))
VOUCHER_GRANT           = int(os.getenv("VOUCHER_GRANT", 50))

def _load_voucher_codes():
    """Return a set of valid voucher codes (upper-cased) from env."""
    raw = os.getenv("VOUCHER_CODES", "")
    return {c.strip().upper() for c in raw.split(",") if c.strip()}


# ── Magic byte signatures ───────────────────────────────────────────────────
MAGIC_BYTES = {
    ".pdf":  [b"%PDF"],
    ".docx": [b"PK\x03\x04"],
    ".doc":  [b"\xd0\xcf\x11\xe0"],
    ".xlsx": [b"PK\x03\x04"],
    ".xls":  [b"\xd0\xcf\x11\xe0"],
}


def _paypal_access_token():
    resp = _requests.post(
        f"{PAYPAL_API_BASE}/v1/oauth2/token",
        data={"grant_type": "client_credentials"},
        auth=(PAYPAL_CLIENT_ID, PAYPAL_CLIENT_SECRET),
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]

# ── CSRF ───────────────────────────────────────────────────────────────────
csrf = CSRFProtect(app)

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    logger.warning("CSRF validation failed: %s", e.description)
    payload = {"error": "CSRF token validation failed. Please refresh and try again."}
    if app.debug or os.getenv("CSRF_ERROR_DETAILS", "False").lower() == "true":
        payload["details"] = e.description
    return jsonify(payload), 400

# ── CSP nonce + security headers ───────────────────────────────────────────
@app.before_request
def set_csp_nonce():
    g.csp_nonce = secrets.token_hex(16)

@app.context_processor
def inject_csp_nonce():
    return {"csp_nonce": getattr(g, "csp_nonce", "")}

@app.after_request
def set_security_headers(response):
    nonce = getattr(g, "csp_nonce", "")
    response.headers["X-Frame-Options"]        = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"]        = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"]     = "camera=(), microphone=(), geolocation=()"
    response.headers["Content-Security-Policy"] = (
        f"default-src 'self'; "
        f"script-src 'self' 'nonce-{nonce}' 'unsafe-hashes' "
            f"https://*.paypal.com https://*.paypalobjects.com; "
        f"style-src 'self' 'unsafe-inline' https://www.paypalobjects.com https://fonts.googleapis.com; "
        f"img-src 'self' data: https://*.paypal.com https://*.paypalobjects.com; "
        f"connect-src 'self' https://*.paypal.com https://api.paypal.com; "
        f"frame-src https://*.paypal.com https://www.paypal.com; "
        f"form-action 'self' https://www.paypal.com https://*.paypal.com; "
        f"font-src 'self' https://www.paypalobjects.com https://fonts.googleapis.com https://fonts.gstatic.com; "
        f"base-uri 'self';"
    )
    return response

# ── Rate limiting ──────────────────────────────────────────────────────────
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri=os.getenv("RATELIMIT_STORAGE_URI", "memory://"),
)

# ── Uploads folder ─────────────────────────────────────────────────────────
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# ── Periodic cleanup ───────────────────────────────────────────────────────
_last_cleanup     = 0.0
_CLEANUP_INTERVAL = 300

def cleanup_old_files(max_age_hours=None):
    if max_age_hours is None:
        max_age_hours = app.config["FILE_EXPIRY_HOURS"]
    upload_folder   = Path(app.config["UPLOAD_FOLDER"])
    if not upload_folder.exists():
        return
    current_time    = time.time()
    max_age_seconds = max_age_hours * 3600
    try:
        for file_path in upload_folder.glob("*"):
            if file_path.is_file() and (current_time - file_path.stat().st_mtime) > max_age_seconds:
                file_path.unlink()
                logger.info("Cleaned up old file: %s", file_path.name)
    except Exception as exc:
        logger.error("Error during file cleanup: %s", exc)

@app.before_request
def cleanup_before_request():
    global _last_cleanup
    now = time.time()
    if now - _last_cleanup >= _CLEANUP_INTERVAL:
        _last_cleanup = now
        cleanup_old_files()

# ── Conversion helpers ─────────────────────────────────────────────────────

def _run_with_timeout(fn, args, timeout_seconds):
    result    = [None]
    exception = [None]

    def target():
        try:
            result[0] = fn(*args)
        except Exception as exc:
            exception[0] = exc

    t = threading.Thread(target=target, daemon=True)
    t.start()
    t.join(timeout_seconds)

    if t.is_alive():
        raise TimeoutError(f"Conversion exceeded {timeout_seconds}s time limit.")
    if exception[0]:
        raise exception[0]
    return result[0]


def pdf_to_word(src, out):
    from pdf2docx import Converter
    cv = Converter(src)
    cv.convert(out)
    cv.close()


def pdf_to_excel(src, out):
    import pdfplumber, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    with pdfplumber.open(src) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        ws.append([c if c else "" for c in row])
            else:
                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        ws.append([line])
    wb.save(out)


def word_to_pdf(src, out):
    """Convert .docx/.doc to PDF using LibreOffice headless (preserves full layout)."""
    import subprocess, tempfile

    if not LIBREOFFICE_PATH:
        raise RuntimeError(
            "Word to PDF conversion requires LibreOffice which is not installed on this server."
        )

    with tempfile.TemporaryDirectory() as tmp_dir:
        result = subprocess.run(
            [
                LIBREOFFICE_PATH,
                "--headless",
                "--norestore",
                "--convert-to", "pdf",
                "--outdir", tmp_dir,
                src,
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )

        if result.returncode != 0:
            logger.error("LibreOffice stderr: %s", result.stderr)
            raise Exception(f"LibreOffice conversion failed: {result.stderr.strip()[:200]}")

        base_name = os.path.splitext(os.path.basename(src))[0]
        lo_output = os.path.join(tmp_dir, f"{base_name}.pdf")

        if not os.path.exists(lo_output):
            raise Exception("LibreOffice ran but produced no output file")

        shutil.move(lo_output, out)

    logger.info("word_to_pdf: converted via LibreOffice headless")


def excel_to_pdf(src, out):
    import openpyxl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else "" for cell in row])

    if not data:
        doc = SimpleDocTemplate(out, pagesize=A4)
        doc.build([])
        return

    col_count = len(data[0])
    page_size = landscape(A4) if col_count > 6 else A4
    page_width = page_size[0] - 1.0 * inch
    col_width  = min(1.5 * inch, page_width / col_count) if col_count else 1.5 * inch

    doc   = SimpleDocTemplate(out, pagesize=page_size,
                               topMargin=0.5*inch, bottomMargin=0.5*inch,
                               leftMargin=0.5*inch, rightMargin=0.5*inch)
    table = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#4f6ef7")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  9),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  10),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f2ff")]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#c5caff")),
    ]))
    doc.build([table])


# ── Conversion modes ───────────────────────────────────────────────────────
MODES = {
    "pdf-to-word":  {"fn": pdf_to_word,  "ext": ".docx", "input_ext": [".pdf"]},
    "pdf-to-excel": {"fn": pdf_to_excel, "ext": ".xlsx", "input_ext": [".pdf"]},
    "word-to-pdf":  {"fn": word_to_pdf,  "ext": ".pdf",  "input_ext": [".docx", ".doc"]},
    "excel-to-pdf": {"fn": excel_to_pdf, "ext": ".pdf",  "input_ext": [".xlsx", ".xls"]},
}

ALLOWED_MIME_TYPES = {
    ".pdf":  ["application/pdf"],
    ".docx": ["application/vnd.openxmlformats-officedocument.wordprocessingml.document",
              "application/octet-stream", "application/zip"],
    ".doc":  ["application/msword", "application/octet-stream"],
    ".xlsx": ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
              "application/octet-stream", "application/zip"],
    ".xls":  ["application/vnd.ms-excel", "application/octet-stream"],
}

# ── File validation ────────────────────────────────────────────────────────
def validate_file(file, allowed_extensions, max_size):
    if not file or file.filename == "":
        return False, "No file provided."
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        return False, f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
    if hasattr(file, "content_type"):
        allowed_mimes = ALLOWED_MIME_TYPES.get(file_ext, [])
        if allowed_mimes and file.content_type not in allowed_mimes:
            return False, f"Invalid file format for {file_ext} file."
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    if file_size > max_size:
        return False, f"File too large. Maximum size is {max_size // (1024*1024)} MB."
    if file_size == 0:
        return False, "File is empty."
    return True, None

# ── Routes ─────────────────────────────────────────────────────────────────

# Tool metadata for individual pages (SEO, categories, descriptions)
TOOL_METADATA = {
    "pdf-to-word":     {"title": "Convert PDF to Word Free — Convertly", "desc": "Free online PDF to Word converter. Convert PDFs to editable Word documents instantly. No sign-up required.", "category": "Convert", "emoji": "📄"},
    "word-to-pdf":     {"title": "Convert Word to PDF Free — Convertly", "desc": "Free Word to PDF converter. Convert .docx and .doc files to PDF online instantly, no registration needed.", "category": "Convert", "emoji": "📝"},
    "pdf-to-excel":    {"title": "Convert PDF to Excel Free — Convertly", "desc": "Extract tables from PDFs to Excel. Free online PDF to Excel converter. Convert instantly, no sign-up.", "category": "Convert", "emoji": "📊"},
    "excel-to-pdf":    {"title": "Convert Excel to PDF Free — Convertly", "desc": "Free Excel to PDF converter. Convert spreadsheets to PDF online instantly. No registration required.", "category": "Convert", "emoji": "📈"},
    "pdf-to-jpg":      {"title": "Convert PDF to JPG Free — Convertly", "desc": "Free PDF to JPG converter. Convert PDF pages to images instantly. Perfect for sharing and viewing.", "category": "Convert", "emoji": "🖼️"},
    "jpg-to-pdf":      {"title": "Convert JPG to PDF Free — Convertly", "desc": "Free JPG to PDF converter. Turn images into a single PDF file instantly. No sign-up needed.", "category": "Convert", "emoji": "🗃️"},
    "merge-pdf":       {"title": "Merge PDF Free — Convertly", "desc": "Free PDF merger. Combine multiple PDF files into one instantly. No registration, no ads.", "category": "Organise", "emoji": "📎"},
    "split-pdf":       {"title": "Split PDF Free — Convertly", "desc": "Free online PDF splitter. Extract specific pages or page ranges from PDF files instantly.", "category": "Organise", "emoji": "✂️"},
    "compress-pdf":    {"title": "Compress PDF Free — Convertly", "desc": "Free PDF compressor. Reduce PDF file size instantly while maintaining quality. No sign-up needed.", "category": "Organise", "emoji": "🗜️"},
    "remove-pages":    {"title": "Remove Pages from PDF Free — Convertly", "desc": "Free PDF page remover. Delete unwanted pages from your PDF instantly. No registration required.", "category": "Organise", "emoji": "🗑️"},
    "extract-pages":   {"title": "Extract PDF Pages Free — Convertly", "desc": "Free PDF page extractor. Keep only the pages you need from your PDF file. Instant, no sign-up.", "category": "Organise", "emoji": "📤"},
    "organize-pdf":    {"title": "Organize PDF Pages Free — Convertly", "desc": "Free PDF organizer. Drag and drop to reorder pages in your PDF. Instant results, no registration.", "category": "Organise", "emoji": "📐"},
    "rotate-pdf":      {"title": "Rotate PDF Pages Free — Convertly", "desc": "Free PDF rotator. Rotate PDF pages any direction instantly. Perfect for fixing orientation issues.", "category": "Organise", "emoji": "🔄"},
    "page-numbers":    {"title": "Add Page Numbers to PDF Free — Convertly", "desc": "Free page number stamper. Add numbers to every page of your PDF. Quick, easy, no sign-up.", "category": "Organise", "emoji": "🔢"},
    "watermark-pdf":   {"title": "Add Watermark to PDF Free — Convertly", "desc": "Free PDF watermark tool. Stamp text on every page of your PDF. No registration required.", "category": "Edit & Sign", "emoji": "💧"},
    "edit-pdf":        {"title": "Edit PDF Free — Convertly", "desc": "Free online PDF editor. Add and edit text on your PDF pages instantly. No subscription needed.", "category": "Edit & Sign", "emoji": "✏️"},
    "sign-pdf":        {"title": "Sign PDF Free — Convertly", "desc": "Free PDF signature tool. Draw, type, or upload your signature to PDF files. Instant, no registration.", "category": "Edit & Sign", "emoji": "✍️"},
    "translate-pdf":   {"title": "Translate PDF Free — AI Powered — Convertly", "desc": "Free AI-powered PDF translator. Translate PDFs into 23 languages instantly. No sign-up required.", "category": "Edit & Sign", "emoji": "🌐"},
    "unlock-pdf":      {"title": "Unlock PDF Free — Convertly", "desc": "Free PDF password remover. Remove password protection from PDF files instantly. No sign-up needed.", "category": "Protect", "emoji": "🔓"},
    "protect-pdf":     {"title": "Protect PDF with Password Free — Convertly", "desc": "Free PDF password protector. Add password protection to your PDF files instantly. Secure & easy.", "category": "Protect", "emoji": "🔐"},
}

@app.route("/")
def index():
    """Homepage with homepage SEO metadata."""
    return render_template("index.html", 
        tool=None,
        page_title="Convertly — Free PDF Converter Online",
        page_description="Free PDF converter online. Convert PDF to Word, Excel, JPG, and more. Translate PDFs in 23 languages. Merge, split, and compress PDFs. No sign-up needed.",
        og_title="Convertly — Free PDF Converter Online",
        og_description="Convert PDFs, translate to 23 languages, merge, split, compress. Free, instant, no registration."
    )

@app.route("/test-reset-quota", methods=["POST"])
def test_reset_quota():
    """Reset conversion quota for testing. Requires secret key."""
    secret = request.form.get("secret") or request.headers.get("X-Test-Secret")
    expected = os.getenv("TEST_SECRET", "")
    if not expected or secret != expected:
        return jsonify({"error": "Unauthorized"}), 401
    fingerprint = get_client_fingerprint(request)
    ConversionCounter.reset(fingerprint)
    ConversionCounter.grant_pro(fingerprint, 9999)
    return jsonify({"reset": True, "pro": True})

@app.route("/<tool_slug>")
def tool_page(tool_slug):
    """Individual tool page with unique SEO metadata."""
    # Normalize the slug (convert underscores to hyphens if needed)
    tool_slug = tool_slug.lower().replace("_", "-")
    
    if tool_slug not in TOOL_METADATA:
        return redirect("/")
    
    meta = TOOL_METADATA[tool_slug]
    return render_template("index.html",
        tool=tool_slug,
        page_title=meta["title"],
        page_description=meta["desc"],
        og_title=meta["title"],
        og_description=meta["desc"]
    )


@app.route("/share")
def share():
    return render_template("share.html")


@app.route("/status")
def status():
    fingerprint = get_client_fingerprint(request)
    used, budget, remaining, is_pro = ConversionCounter.get_status(fingerprint)
    
    # Check JWT Pro token
    pro_token = request.cookies.get("pro_token")
    pro_valid = False
    if pro_token:
        valid, payload = ProToken.verify(pro_token)
        if valid and payload.get("fingerprint") == fingerprint:
            pro_valid = True
    
    is_pro = is_pro or pro_valid
    
    return jsonify({
        "conversions_used":      used,
        "conversions_budget":    budget,
        "conversions_remaining": remaining,
        "paid":                  is_pro,
        "pro_valid":             is_pro,
        "free_limit":            FREE_CONVERSIONS_LIMIT,
        "paid_amount":           PAID_CONVERSIONS_AMOUNT,
        "price_usd":             PAYPAL_PRICE_USD,
    })


@app.route("/google4bbdce78d1b4593c.html")
def google_verification():
    return "google-site-verification: google4bbdce78d1b4593c.html"


@app.route("/convert", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "20 per minute; 200 per hour"))
def convert():
    # ── Validate mode & file ───────────────────────────────────────────────
    mode = request.form.get("mode")
    file = request.files.get("file")

    if not mode or mode not in MODES:
        return jsonify({"error": "Invalid conversion mode."}), 400

    allowed_exts = MODES[mode]["input_ext"]
    is_valid, error_msg = validate_file(file, allowed_exts, app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    file_ext = os.path.splitext(file.filename)[1].lower()
    expected_magic = MAGIC_BYTES.get(file_ext, [])
    if expected_magic:
        header = file.read(8)
        file.seek(0)
        if not any(header.startswith(m) for m in expected_magic):
            logger.warning("Magic byte mismatch for file: %s", file.filename)
            return jsonify({"error": "File content does not match its extension."}), 400

    # ── SERVER-SIDE Quota check using fingerprint ──────────────────────────
    fingerprint = get_client_fingerprint(request)
    used, budget, remaining, is_pro = ConversionCounter.get_status(fingerprint)

    if used >= budget and not is_pro:
        return jsonify({
            "error": "quota_exceeded",
            "message": "You've used all your free conversions.",
            "used": used,
            "budget": budget,
        }), 402

    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out = os.path.splitext(src)[0] + MODES[mode]["ext"]
    file.save(src)
    logger.info("Converting [%s] %s", mode, safe_name)

    try:
        _run_with_timeout(MODES[mode]["fn"], (src, out), app.config["CONVERSION_TIMEOUT"])
    except TimeoutError as exc:
        logger.error("Conversion timeout for %s: %s", safe_name, exc)
        return jsonify({"error": str(exc)}), 504
    except Exception as exc:
        logger.error("Conversion error for %s: %s", safe_name, exc, exc_info=True)
        if isinstance(exc, RuntimeError) and "not installed" in str(exc):
            return jsonify({"error": "Word to PDF conversion is temporarily unavailable. Please try again later."}), 503
        return jsonify({"error": "Conversion failed. Please check your file and try again."}), 500
    finally:
        if os.path.exists(src):
            os.remove(src)

    if not os.path.exists(out):
        logger.error("Output file missing after conversion: %s", out)
        return jsonify({"error": "Conversion produced no output. Please try again."}), 500

    # ── Increment server-side quota counter ────────────────────────────────
    used, budget, remaining = ConversionCounter.increment(fingerprint, request)

    out_name = os.path.splitext(safe_name)[0] + "_converted" + MODES[mode]["ext"]
    logger.info("Conversion complete: %s → %s", safe_name, out_name)

    response = send_file(out, as_attachment=True, download_name=out_name)
    
    # Add quota headers to response
    response.headers["X-Conversions-Used"] = str(used)
    response.headers["X-Conversions-Remaining"] = str(remaining)

    @after_this_request
    def remove_output(resp):
        try:
            if os.path.exists(out):
                os.remove(out)
        except Exception:
            pass
        return resp

    return response


# ── Merge PDF ──────────────────────────────────────────────────────────────

@app.route("/merge-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def merge_pdf_route():
    files = request.files.getlist("files[]")
    if len(files) < 2:
        return jsonify({"error": "Please upload at least 2 PDF files to merge."}), 400
    if len(files) > 20:
        return jsonify({"error": "Maximum 20 files can be merged at once."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Validate every file before saving anything ─────────────────────────
    total_size = 0
    for f in files:
        if not f or f.filename == "":
            return jsonify({"error": "One or more files is missing or has no name."}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext != ".pdf":
            return jsonify({"error": f"All files must be PDFs. '{os.path.basename(f.filename)}' is not a PDF."}), 400
        f.seek(0, 2)
        size = f.tell()
        f.seek(0)
        if size == 0:
            return jsonify({"error": f"File '{os.path.basename(f.filename)}' is empty."}), 400
        header = f.read(8)
        f.seek(0)
        if not header.startswith(b"%PDF"):
            return jsonify({"error": f"'{os.path.basename(f.filename)}' does not appear to be a valid PDF."}), 400
        total_size += size

    max_total = app.config["MAX_FILE_SIZE"] * 3   # up to 3× the single-file cap
    if total_size > max_total:
        return jsonify({"error": f"Total size too large. Maximum combined size is {max_total // (1024 * 1024)} MB."}), 400

    uid         = str(uuid.uuid4())
    saved_paths = []
    out_path    = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_merged.pdf")

    try:
        # Save every uploaded file to a temp location
        for i, f in enumerate(files):
            safe_name = os.path.basename(f.filename)
            tmp_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{i}_{safe_name}")
            f.save(tmp_path)
            saved_paths.append(tmp_path)

        # Merge with PyMuPDF (already a dependency)
        import fitz  # PyMuPDF
        merged_doc = fitz.open()
        for path in saved_paths:
            src_doc = fitz.open(path)
            merged_doc.insert_pdf(src_doc)
            src_doc.close()
        merged_doc.save(out_path)
        merged_doc.close()
        logger.info("Merge PDF: merged %d files → %s", len(saved_paths), out_path)

    except Exception as exc:
        logger.error("Merge PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to merge PDFs. Please check that all files are valid, non-corrupted PDFs."}), 500
    finally:
        # Always remove the source temp files
        for path in saved_paths:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass

    if not os.path.exists(out_path):
        logger.error("Merge output file missing: %s", out_path)
        return jsonify({"error": "Merge produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    @after_this_request
    def remove_merged(response):
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
        return response

    return send_file(out_path, as_attachment=True, download_name="merged.pdf")


# ── Split PDF helpers ──────────────────────────────────────────────────────

def _parse_page_ranges(ranges_str, total_pages):
    """Parse '1-3, 5, 7-9' into a list of (start_0idx, end_0idx) tuples.
    Raises ValueError with a human-readable message on any invalid input."""
    result = []
    for part in ranges_str.replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            sides = part.split("-", 1)
            try:
                start = int(sides[0].strip())
                end   = int(sides[1].strip())
            except ValueError:
                raise ValueError(f"Invalid range: '{part}'. Use numbers like '1-3'.")
            if start < 1 or end < 1:
                raise ValueError(f"Page numbers must be 1 or greater (got '{part}').")
            if start > total_pages or end > total_pages:
                raise ValueError(
                    f"Page {max(start, end)} is out of range. "
                    f"This PDF has {total_pages} page{'s' if total_pages != 1 else ''}."
                )
            if start > end:
                raise ValueError(f"Range '{part}' is invalid — start must be <= end.")
            result.append((start - 1, end - 1))
        else:
            try:
                page = int(part)
            except ValueError:
                raise ValueError(f"Invalid page number: '{part}'.")
            if page < 1:
                raise ValueError(f"Page numbers must be 1 or greater (got '{page}').")
            if page > total_pages:
                raise ValueError(
                    f"Page {page} is out of range. "
                    f"This PDF has {total_pages} page{'s' if total_pages != 1 else ''}."
                )
            result.append((page - 1, page - 1))
    return result


# ── Split PDF route ────────────────────────────────────────────────────────

@app.route("/split-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def split_pdf_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Validate split parameters ──────────────────────────────────────────
    split_mode = request.form.get("split_mode", "ranges")
    if split_mode not in ("ranges", "every_n", "individual"):
        return jsonify({"error": "Invalid split mode."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Save source file ───────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    file.save(src_path)

    output_paths = []   # list of (abs_file_path, archive_name)
    try:
        import fitz   # PyMuPDF
        src_doc     = fitz.open(src_path)
        total_pages = src_doc.page_count

        if total_pages == 0:
            return jsonify({"error": "PDF has no pages."}), 400

        # ── Build segment list: [(from_0idx, to_0idx, archive_name)] ───────
        segments = []

        if split_mode == "individual":
            if total_pages > 200:
                return jsonify({
                    "error": f"PDF has {total_pages} pages. Individual split is capped at 200 pages."
                }), 400
            for i in range(total_pages):
                segments.append((i, i, f"page_{i + 1}.pdf"))

        elif split_mode == "every_n":
            try:
                n = int(request.form.get("n_pages", 1))
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid value for N pages."}), 400
            if n < 1:
                return jsonify({"error": "N pages must be at least 1."}), 400
            if n >= total_pages:
                return jsonify({
                    "error": f"N ({n}) must be less than the total pages ({total_pages})."
                }), 400
            i = 0
            while i < total_pages:
                end = min(i + n - 1, total_pages - 1)
                name = f"page_{i + 1}.pdf" if i == end else f"pages_{i + 1}-{end + 1}.pdf"
                segments.append((i, end, name))
                i += n

        else:   # split_mode == "ranges"
            ranges_str = request.form.get("ranges", "").strip()
            if not ranges_str:
                return jsonify({"error": "Please enter at least one page range."}), 400
            try:
                parsed = _parse_page_ranges(ranges_str, total_pages)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
            if not parsed:
                return jsonify({"error": "No valid page ranges found."}), 400
            for start, end in parsed:
                name = f"page_{start + 1}.pdf" if start == end else f"pages_{start + 1}-{end + 1}.pdf"
                segments.append((start, end, name))

        if not segments:
            return jsonify({"error": "No pages to extract."}), 400

        # Deduplicate archive names in case of overlapping/identical ranges
        seen: dict = {}
        final_segments = []
        for start, end, name in segments:
            base, ext = os.path.splitext(name)
            if name in seen:
                seen[name] += 1
                name = f"{base}_{seen[name]}{ext}"
            else:
                seen[name] = 0
            final_segments.append((start, end, name))

        # ── Create output PDFs ─────────────────────────────────────────────
        for start, end, name in final_segments:
            out_doc  = fitz.open()
            out_doc.insert_pdf(src_doc, from_page=start, to_page=end)
            out_file = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_split_{name}")
            out_doc.save(out_file)
            out_doc.close()
            output_paths.append((out_file, name))

        src_doc.close()

    except Exception as exc:
        logger.error("Split PDF error: %s", exc, exc_info=True)
        for path, _ in output_paths:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass
        return jsonify({"error": "Failed to split PDF. Please check your file and parameters."}), 500
    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    if not output_paths:
        return jsonify({"error": "Split produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)
    logger.info("Split PDF: %d segment(s) from %s", len(output_paths), safe_name)

    base_name = os.path.splitext(safe_name)[0]

    # ── Single output → return as a plain PDF ─────────────────────────────
    if len(output_paths) == 1:
        single_path, single_name = output_paths[0]
        download_name = f"{base_name}_{single_name}"

        @after_this_request
        def remove_single(response):
            try:
                if os.path.exists(single_path):
                    os.remove(single_path)
            except Exception:
                pass
            return response

        return send_file(single_path, as_attachment=True, download_name=download_name)

    # ── Multiple outputs → ZIP ─────────────────────────────────────────────
    import io as _io, zipfile as _zipfile

    zip_buf = _io.BytesIO()
    try:
        with _zipfile.ZipFile(zip_buf, "w", _zipfile.ZIP_DEFLATED) as zf:
            for out_file, name in output_paths:
                zf.write(out_file, name)
    finally:
        for out_file, _ in output_paths:
            if os.path.exists(out_file):
                try:
                    os.remove(out_file)
                except Exception:
                    pass

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"{base_name}_split.zip",
    )


# ── Remove Pages route ─────────────────────────────────────────────────────

@app.route("/remove-pages", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def remove_pages_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Validate pages input ───────────────────────────────────────────────
    pages_input = request.form.get("pages", "").strip()
    if not pages_input:
        return jsonify({"error": "Please enter at least one page number to remove."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Save source file ───────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_removed.pdf")
    file.save(src_path)

    try:
        import fitz   # PyMuPDF
        src_doc     = fitz.open(src_path)
        total_pages = src_doc.page_count

        if total_pages == 0:
            return jsonify({"error": "PDF has no pages."}), 400

        # Parse the page ranges into a set of 0-based indices to remove
        try:
            parsed = _parse_page_ranges(pages_input, total_pages)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        pages_to_remove = set()
        for start, end in parsed:
            for i in range(start, end + 1):
                pages_to_remove.add(i)

        if not pages_to_remove:
            return jsonify({"error": "No valid pages specified for removal."}), 400

        if len(pages_to_remove) >= total_pages:
            return jsonify({
                "error": f"Cannot remove all {total_pages} pages. "
                         f"The result must contain at least 1 page."
            }), 400

        # Build a new document from the pages we want to KEEP.
        # This avoids mutating + saving the source doc (which causes Windows
        # file-lock issues with PyMuPDF's incremental-save internals).
        new_doc = fitz.open()
        for i in range(total_pages):
            if i not in pages_to_remove:
                new_doc.insert_pdf(src_doc, from_page=i, to_page=i)
        src_doc.close()
        new_doc.save(out_path, garbage=4, deflate=True)
        new_doc.close()

        remaining = total_pages - len(pages_to_remove)
        logger.info(
            "Remove pages: removed %d page(s) from %s (%d remaining)",
            len(pages_to_remove), safe_name, remaining,
        )

    except Exception as exc:
        logger.error("Remove pages error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to remove pages. Please check your file and page numbers."}), 500
    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Operation produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_removed.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
        return response

    return send_file(out_path, as_attachment=True, download_name=download_name)


# ── Extract Pages route ────────────────────────────────────────────────────

@app.route("/extract-pages", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def extract_pages_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Validate pages input ───────────────────────────────────────────────
    pages_input = request.form.get("pages", "").strip()
    if not pages_input:
        return jsonify({"error": "Please enter at least one page number or range to extract."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Save source file ───────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_extracted.pdf")
    file.save(src_path)

    try:
        import fitz   # PyMuPDF
        src_doc     = fitz.open(src_path)
        total_pages = src_doc.page_count

        if total_pages == 0:
            return jsonify({"error": "PDF has no pages."}), 400

        try:
            parsed = _parse_page_ranges(pages_input, total_pages)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        if not parsed:
            return jsonify({"error": "No valid page ranges found."}), 400

        # Build output doc by inserting each requested range in the order given.
        # This also allows users to reorder pages (e.g. "5, 1-3" puts p.5 first).
        new_doc = fitz.open()
        for start, end in parsed:
            new_doc.insert_pdf(src_doc, from_page=start, to_page=end)
        src_doc.close()

        if new_doc.page_count == 0:
            new_doc.close()
            return jsonify({"error": "Extraction produced an empty document."}), 400

        new_doc.save(out_path, garbage=4, deflate=True)
        new_doc.close()

        logger.info("Extract pages: %d segment(s) → %d page(s) from %s",
                    len(parsed), new_doc.page_count if not new_doc.is_closed else "?", safe_name)

    except Exception as exc:
        logger.error("Extract pages error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to extract pages. Please check your file and page numbers."}), 500
    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Extraction produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_extracted.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
        return response

    return send_file(out_path, as_attachment=True, download_name=download_name)


# ── Compress PDF helpers ───────────────────────────────────────────────────

def _compress_pdf_lossless(src_path, out_path):
    """Re-save the PDF with full structural cleanup and stream deflation.
    No quality loss — works best on text/vector PDFs."""
    import fitz
    doc = fitz.open(src_path)
    doc.save(
        out_path,
        garbage=4,            # purge all unused objects, streams, and duplicates
        deflate=True,         # compress all uncompressed streams
        deflate_images=True,  # compress image streams
        deflate_fonts=True,   # compress font streams
        clean=True,           # sanitise content streams
        linear=False,
    )
    doc.close()


def _compress_pdf_render(src_path, out_path, dpi, jpeg_quality):
    """Re-render every page as a JPEG image at `dpi` resolution.
    Produces a flat image-based PDF — text will NOT be selectable after this.
    Processes one page at a time to keep peak RAM low."""
    import fitz
    src_doc = fitz.open(src_path)
    new_doc = fitz.open()
    scale   = dpi / 72.0
    mat     = fitz.Matrix(scale, scale)

    for i in range(src_doc.page_count):
        page     = src_doc.load_page(i)
        pix      = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
        img_data = pix.tobytes("jpeg", jpg_quality=jpeg_quality)
        del pix                            # free pixmap RAM immediately

        new_page = new_doc.new_page(width=page.rect.width, height=page.rect.height)
        new_page.insert_image(new_page.rect, stream=img_data)
        del img_data

    src_doc.close()
    new_doc.save(out_path, garbage=4, deflate=True)
    new_doc.close()


# ── Compress PDF route ─────────────────────────────────────────────────────

@app.route("/compress-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def compress_pdf_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Validate compression level ─────────────────────────────────────────
    level = request.form.get("level", "lossless")
    if level not in ("lossless", "balanced", "maximum"):
        return jsonify({"error": "Invalid compression level."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Save source file ───────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_compressed.pdf")
    file.save(src_path)

    original_size = os.path.getsize(src_path)

    LEVEL_CONFIG = {
        "lossless": {"fn": _compress_pdf_lossless, "args": []},
        "balanced": {"fn": _compress_pdf_render,   "args": [150, 80]},
        "maximum":  {"fn": _compress_pdf_render,   "args": [96,  65]},
    }
    cfg = LEVEL_CONFIG[level]

    try:
        _run_with_timeout(
            cfg["fn"],
            (src_path, out_path, *cfg["args"]),
            app.config["CONVERSION_TIMEOUT"],
        )
    except TimeoutError as exc:
        logger.error("Compress PDF timeout: %s", exc)
        return jsonify({"error": str(exc)}), 504
    except Exception as exc:
        logger.error("Compress PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to compress PDF. Please check your file and try again."}), 500
    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Compression produced no output. Please try again."}), 500

    compressed_size = os.path.getsize(out_path)
    pct = round((1 - compressed_size / original_size) * 100, 1) if original_size > 0 else 0
    logger.info(
        "Compress PDF [%s]: %d B → %d B (%.1f%% reduction), file=%s",
        level, original_size, compressed_size, pct, safe_name,
    )

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_compressed.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=download_name)
    # Surface size stats to the client via custom headers
    response.headers["X-Original-Size"]   = str(original_size)
    response.headers["X-Compressed-Size"] = str(compressed_size)
    response.headers["X-Reduction-Pct"]   = str(pct)
    response.headers["Access-Control-Expose-Headers"] = (
        "X-Original-Size, X-Compressed-Size, X-Reduction-Pct"
    )
    return response


# ── Organize PDF routes ────────────────────────────────────────────────────

@app.route("/organize-pdf/preview", methods=["POST"])
@limiter.limit("10 per minute")
def organize_pdf_preview():
    """Step 1 — accept a PDF, return JPEG thumbnails + page count.
    Does NOT consume a quota slot; only reorder does."""
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_organize_{safe_name}")
    file.save(src_path)

    try:
        import fitz, base64 as _b64

        src_doc     = fitz.open(src_path)
        total_pages = src_doc.page_count

        if total_pages == 0:
            return jsonify({"error": "PDF has no pages."}), 400
        if total_pages > 50:
            return jsonify({
                "error": f"PDF has {total_pages} pages. Organize is limited to 50 pages."
            }), 400

        # Render each page to a small JPEG thumbnail
        mat        = fitz.Matrix(0.18, 0.18)   # ≈ 107×152 px for A4
        thumbnails = []
        for i in range(total_pages):
            page      = src_doc.load_page(i)
            pix       = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            try:
                img_bytes = pix.tobytes("jpeg", jpg_quality=65)
            except Exception:
                img_bytes = pix.tobytes("png")   # fallback
            thumbnails.append(_b64.b64encode(img_bytes).decode())

        src_doc.close()

    except Exception as exc:
        logger.error("Organize PDF preview error: %s", exc, exc_info=True)
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to generate preview. Is the PDF valid and not encrypted?"}), 500

    # Store a reference in the session so only this session can reorder it
    session["organize_preview_id"] = uid
    session["organize_file_name"]  = safe_name
    session.modified = True

    logger.info("Organize PDF preview: %d pages, uid=%s, file=%s", total_pages, uid, safe_name)
    return jsonify({
        "preview_id": uid,
        "page_count": total_pages,
        "file_name":  safe_name,
        "thumbnails": thumbnails,
    })


@app.route("/organize-pdf/reorder", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def organize_pdf_reorder():
    """Step 2 — apply new page order, return reordered PDF. Consumes a quota slot."""
    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Validate preview session ───────────────────────────────────────────
    preview_id     = request.form.get("preview_id", "").strip()
    session_uid    = session.get("organize_preview_id", "")
    safe_name      = session.get("organize_file_name", "organized.pdf")

    if not preview_id or preview_id != session_uid:
        return jsonify({"error": "Invalid or expired preview. Please upload your file again."}), 400

    # Find the saved source file
    src_path = None
    for candidate in Path(app.config["UPLOAD_FOLDER"]).glob(f"{preview_id}_organize_*"):
        src_path = str(candidate)
        break

    if not src_path or not os.path.exists(src_path):
        return jsonify({"error": "Preview session expired. Please upload your file again."}), 400

    # ── Parse new order ────────────────────────────────────────────────────
    order_raw = request.form.getlist("order[]")
    if not order_raw:
        return jsonify({"error": "No page order provided."}), 400
    try:
        new_order = [int(x) for x in order_raw]
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid page order data."}), 400

    out_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{preview_id}_reordered.pdf")

    try:
        import fitz
        src_doc     = fitz.open(src_path)
        total_pages = src_doc.page_count

        if len(new_order) != total_pages:
            return jsonify({"error": f"Order must include all {total_pages} pages."}), 400
        if sorted(new_order) != list(range(total_pages)):
            return jsonify({"error": "Invalid order: must be a permutation of all page indices."}), 400

        new_doc = fitz.open()
        for page_idx in new_order:
            new_doc.insert_pdf(src_doc, from_page=page_idx, to_page=page_idx)
        src_doc.close()

        new_doc.save(out_path, garbage=4, deflate=True)
        new_doc.close()

        logger.info("Organize PDF reorder: %d pages reordered, uid=%s", total_pages, preview_id)

    except Exception as exc:
        logger.error("Organize PDF reorder error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to reorder PDF pages. Please try again."}), 500
    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Reorder produced no output. Please try again."}), 500

    # ── Increment quota and clean up session keys ──────────────────────────
    ConversionCounter.increment(fingerprint, request)
    session.pop("organize_preview_id", None)
    session.pop("organize_file_name",  None)
    session.modified = True

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_organized.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
        return response

    return send_file(out_path, as_attachment=True, download_name=download_name)


@app.route("/create-paypal-order", methods=["POST"])
@limiter.limit("5 per hour")
def create_paypal_order():
    if not PAYPAL_CLIENT_ID or not PAYPAL_CLIENT_SECRET:
        # Fall back to direct payment link if API credentials are not set
        return redirect("https://www.paypal.com/ncp/payment/T5VKFKXKW8AYG")

    try:
        token    = _paypal_access_token()
        base_url = request.host_url.rstrip("/")
        resp = _requests.post(
            f"{PAYPAL_API_BASE}/v2/checkout/orders",
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {token}"},
            json={
                "intent": "CAPTURE",
                "purchase_units": [{
                    "amount": {"currency_code": "USD", "value": PAYPAL_PRICE_USD},
                    "description": f"{PAID_CONVERSIONS_AMOUNT} additional file conversions on Convertly",
                }],
                "application_context": {
                    "return_url":  f"{base_url}/payment-success",
                    "cancel_url":  f"{base_url}/?cancelled=1",
                    "brand_name":  "Convertly",
                    "user_action": "PAY_NOW",
                },
            },
            timeout=15,
        )
        resp.raise_for_status()
        order = resp.json()
    except Exception as exc:
        logger.error("PayPal order creation failed: %s", exc)
        return jsonify({"error": "Payment service unavailable. Please try again."}), 503

    approval_url = next(
        (lnk["href"] for lnk in order.get("links", []) if lnk["rel"] == "approve"),
        None,
    )
    if not approval_url:
        return jsonify({"error": "Could not get PayPal approval URL."}), 503

    session["pending_paypal_order_id"] = order["id"]
    session.modified = True
    return redirect(approval_url)


@app.route("/payment-success")
def payment_success():
    # PayPal passes ?token=ORDER_ID&PayerID=... on return
    order_id = request.args.get("token", "")
    payer_id = request.args.get("PayerID", "")

    if not order_id or not payer_id:
        logger.warning("PayPal return missing token or PayerID")
        return redirect("/?error=payment_incomplete")

    try:
        access_token = _paypal_access_token()

        # Step 1: check current order status
        check_resp = _requests.get(
            f"{PAYPAL_API_BASE}/v2/checkout/orders/{order_id}",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=15,
        )
        check_resp.raise_for_status()
        order_data = check_resp.json()
        status = order_data.get("status")

        # Step 2: if APPROVED, capture it — this is what converts approval → payment
        if status == "APPROVED":
            cap_resp = _requests.post(
                f"{PAYPAL_API_BASE}/v2/checkout/orders/{order_id}/capture",
                headers={
                    "Content-Type":  "application/json",
                    "Authorization": f"Bearer {access_token}",
                },
                json={},
                timeout=15,
            )
            cap_resp.raise_for_status()
            order_data = cap_resp.json()
            status = order_data.get("status")

    except Exception as exc:
        logger.error("PayPal capture failed for order %s: %s", order_id, exc)
        return redirect("/?error=payment_error")

    if status != "COMPLETED":
        logger.warning("PayPal order %s status after capture: %s", order_id, status)
        return redirect("/?error=payment_incomplete")

    # ✅ PAYMENT CONFIRMED - Grant Pro access with JWT token
    fingerprint = get_client_fingerprint(request)
    
    # Create signed JWT
    pro_token = ProToken.create(fingerprint, int(os.getenv("PAID_CONVERSIONS_AMOUNT", PAID_CONVERSIONS_AMOUNT)))
    
    # Grant server-side conversions
    ConversionCounter.grant_pro(fingerprint, int(os.getenv("PAID_CONVERSIONS_AMOUNT", PAID_CONVERSIONS_AMOUNT)))
    
    # Store invoice details in session
    unit    = order_data.get("purchase_units", [{}])[0]
    capture = unit.get("payments", {}).get("captures", [{}])[0]
    session["last_invoice"] = {
        "order_id":  order_data.get("id", order_id),
        "item_name": f"{PAID_CONVERSIONS_AMOUNT} Additional File Conversions",
        "price":     capture.get("amount", {}).get("value", PAYPAL_PRICE_USD),
        "currency":  capture.get("amount", {}).get("currency_code", "USD"),
        "date":      capture.get("create_time", ""),
    }
    session.modified = True

    logger.info("PayPal payment COMPLETED for order %s; Pro access granted", order_id)
    
    response = redirect("/invoice")
    # ✅ Set HttpOnly, Secure, SameSite=Strict Pro token
    response.set_cookie(
        "pro_token",
        pro_token,
        httponly=True,
        secure=_is_production,
        samesite="Strict",
        max_age=7*24*3600,  # 7 days
    )
    return response


@app.route("/invoice")
def invoice():
    inv = session.get("last_invoice")
    if not inv:
        return redirect("/")
    from datetime import date
    return render_template("invoice.html", inv=inv, now=date.today().isoformat())


# ── PDF to JPG route ───────────────────────────────────────────────────────

@app.route("/pdf-to-jpg", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def pdf_to_jpg_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── DPI parameter (low=96, medium=150, high=300) ───────────────────────
    DPI_MAP = {"low": 96, "medium": 150, "high": 300}
    dpi = DPI_MAP.get(request.form.get("dpi", "medium"), 150)

    # ── Format parameter (jpg, png, webp) ──────────────────────────────────
    fmt = request.form.get("fmt", "jpg").lower()
    if fmt not in ("jpg", "jpeg", "png", "webp"):
        fmt = "jpg"
    if fmt == "jpeg":
        fmt = "jpg"
    IMG_EXT  = {"jpg": ".jpg", "png": ".png", "webp": ".webp"}[fmt]
    IMG_MIME = {"jpg": "image/jpeg", "png": "image/png", "webp": "image/webp"}[fmt]

    # ── Save source file ───────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    file.save(src_path)

    out_paths = []   # list of (path, filename) tuples to clean up

    try:
        import fitz
        import zipfile
        import io as _io

        doc = fitz.open(src_path)
        total_pages = doc.page_count

        if total_pages == 0:
            doc.close()
            return jsonify({"error": "PDF has no pages."}), 400
        if total_pages > 200:
            doc.close()
            return jsonify({"error": "PDF has too many pages (max 200 for image export)."}), 400

        scale = dpi / 72.0
        mat   = fitz.Matrix(scale, scale)
        base  = os.path.splitext(safe_name)[0]

        if total_pages == 1:
            # ── Single page → return image directly ───────────────────────
            page = doc.load_page(0)
            pix  = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            try:
                if fmt == "jpg":
                    img_bytes = pix.tobytes("jpeg", jpg_quality=92)
                elif fmt == "png":
                    img_bytes = pix.tobytes("png")
                elif fmt == "webp":
                    img_bytes = pix.tobytes("webp", jpg_quality=92)
                else:
                    img_bytes = pix.tobytes("jpeg", jpg_quality=92)
            except Exception:
                img_bytes = pix.tobytes("png")
            del pix
            doc.close()

            out_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{base}_page_1{IMG_EXT}")
            with open(out_path, "wb") as fh:
                fh.write(img_bytes)
            out_paths.append(out_path)

            ConversionCounter.increment(fingerprint, request)

            @after_this_request
            def _remove_single(response):
                for p in out_paths:
                    try:
                        if os.path.exists(p): os.remove(p)
                    except Exception:
                        pass
                return response

            return send_file(
                out_path,
                as_attachment=True,
                download_name=f"{base}_page_1{IMG_EXT}",
                mimetype=IMG_MIME,
            )

        else:
            # ── Multiple pages → build ZIP in memory ──────────────────────
            zip_buf = _io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for i in range(total_pages):
                    page = doc.load_page(i)
                    pix  = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
                    try:
                        if fmt == "jpg":
                            img_bytes = pix.tobytes("jpeg", jpg_quality=92)
                        elif fmt == "png":
                            img_bytes = pix.tobytes("png")
                        elif fmt == "webp":
                            img_bytes = pix.tobytes("webp", jpg_quality=92)
                        else:
                            img_bytes = pix.tobytes("jpeg", jpg_quality=92)
                    except Exception:
                        img_bytes = pix.tobytes("png")
                    del pix
                    fname_in_zip = f"{base}_page_{i + 1}{IMG_EXT}"
                    zf.writestr(fname_in_zip, img_bytes)
            doc.close()

            zip_buf.seek(0)
            zip_name = f"pages_{fmt}.zip"
            zip_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{base}_{zip_name}")
            with open(zip_path, "wb") as fh:
                fh.write(zip_buf.read())
            out_paths.append(zip_path)

            ConversionCounter.increment(fingerprint, request)

            @after_this_request
            def _remove_zip(response):
                for p in out_paths:
                    try:
                        if os.path.exists(p): os.remove(p)
                    except Exception:
                        pass
                return response

            return send_file(
                zip_path,
                as_attachment=True,
                download_name=f"{base}_{zip_name}",
                mimetype="application/zip",
            )

    except Exception as exc:
        logger.error("PDF to JPG error: %s", exc, exc_info=True)
        for p in out_paths:
            try:
                if os.path.exists(p): os.remove(p)
            except Exception:
                pass
        return jsonify({"error": "Could not convert PDF to images. The file may be corrupted or encrypted."}), 500

    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass


# ── JPG / Image to PDF route ───────────────────────────────────────────────

@app.route("/jpg-to-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def jpg_to_pdf_route():
    # ── Validate files ─────────────────────────────────────────────────────
    files = request.files.getlist("files[]")
    files = [f for f in files if f and f.filename]
    if not files:
        return jsonify({"error": "No image files uploaded."}), 400

    MAX_FILES = 30
    if len(files) > MAX_FILES:
        return jsonify({"error": f"Too many files. Maximum is {MAX_FILES} images."}), 400

    ALLOWED_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".tiff", ".tif"}
    for f in files:
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXTS:
            return jsonify({"error": f"'{f.filename}' is not a supported image type. Use JPG, PNG, WEBP, BMP, GIF, or TIFF."}), 400
        f.seek(0, 2)
        size = f.tell()
        f.seek(0)
        if size == 0:
            return jsonify({"error": f"'{f.filename}' is empty."}), 400
        if size > app.config["MAX_FILE_SIZE"]:
            return jsonify({"error": f"'{f.filename}' exceeds the {app.config['MAX_FILE_SIZE'] // (1024*1024)} MB limit."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Save and convert ───────────────────────────────────────────────────
    uid        = str(uuid.uuid4())
    out_path   = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_converted.pdf")
    saved_imgs = []   # (path, delete_after) pairs

    try:
        import fitz
        from PIL import Image as PILImage
        import io as _io

        new_doc = fitz.open()

        for f in files:
            safe_name = os.path.basename(f.filename)
            img_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
            f.save(img_path)
            saved_imgs.append(img_path)

            # Read with Pillow to get exact pixel dimensions and convert exotic
            # formats (WEBP, BMP, GIF, RGBA, palette) to JPEG bytes for MuPDF.
            raw = open(img_path, "rb").read()
            try:
                pil = PILImage.open(_io.BytesIO(raw))
                # Animated GIF — use first frame only
                if hasattr(pil, "n_frames") and pil.n_frames > 1:
                    pil.seek(0)
                # Palette mode: convert to RGBA first to preserve any transparency
                if pil.mode == "P":
                    pil = pil.convert("RGBA")
                # Flatten transparency onto white background
                if pil.mode in ("RGBA", "LA"):
                    bg = PILImage.new("RGB", pil.size, (255, 255, 255))
                    alpha = pil.split()[-1]  # last channel is alpha for both RGBA and LA
                    bg.paste(pil.convert("RGB"), mask=alpha)
                    pil = bg
                elif pil.mode != "RGB":
                    pil = pil.convert("RGB")
                w_px, h_px = pil.size
                buf = _io.BytesIO()
                pil.save(buf, format="JPEG", quality=92)
                embed_bytes = buf.getvalue()
                pil.close()
            except Exception as img_err:
                logger.warning("Pillow failed on %s: %s — using raw bytes", safe_name, img_err)
                embed_bytes = raw
                # Fall back: let MuPDF guess dimensions (best effort)
                w_px, h_px = 800, 600

            # ── Page sizing ────────────────────────────────────────────────
            # Scale so the shorter dimension = 595 pt (A4 short side), preserving
            # aspect ratio.  This gives approx-A4 pages regardless of resolution.
            MIN_SIDE = 595.0
            if w_px <= h_px:
                scale  = MIN_SIDE / max(w_px, 1)
            else:
                scale  = MIN_SIDE / max(h_px, 1)
            page_w = round(w_px * scale)
            page_h = round(h_px * scale)

            page = new_doc.new_page(width=page_w, height=page_h)
            page.insert_image(page.rect, stream=embed_bytes)
            del embed_bytes

        new_doc.save(out_path, garbage=4, deflate=True, deflate_images=True)
        new_doc.close()

        logger.info("JPG→PDF: %d image(s) converted, uid=%s", len(files), uid)

    except Exception as exc:
        logger.error("JPG to PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to convert images to PDF. One or more images may be corrupted."}), 500

    finally:
        for p in saved_imgs:
            try:
                if os.path.exists(p): os.remove(p)
            except Exception:
                pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Conversion produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    page_count    = len(files)
    download_name = "images_converted.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception:
            pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=download_name)
    response.headers["X-Page-Count"] = str(page_count)
    response.headers["Access-Control-Expose-Headers"] = "X-Page-Count"
    return response


# ── Watermark PDF route ────────────────────────────────────────────────────

@app.route("/watermark-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def watermark_pdf_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Watermark parameters ───────────────────────────────────────────────
    text = request.form.get("text", "CONFIDENTIAL").strip()
    if not text:
        return jsonify({"error": "Watermark text cannot be empty."}), 400
    if len(text) > 80:
        return jsonify({"error": "Watermark text is too long (max 80 characters)."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # Opacity: 0.05 – 0.60
    try:
        opacity = float(request.form.get("opacity", "0.25"))
        opacity = max(0.05, min(0.60, opacity))
    except ValueError:
        opacity = 0.25

    # Colour: hex string like "ff0000" → (r, g, b) floats 0-1
    raw_color = request.form.get("color", "808080").lstrip("#")
    try:
        r = int(raw_color[0:2], 16) / 255.0
        g = int(raw_color[2:4], 16) / 255.0
        b = int(raw_color[4:6], 16) / 255.0
    except Exception:
        r, g, b = 0.5, 0.5, 0.5

    # Position: "diagonal" (default), "center", "top", "bottom"
    position = request.form.get("position", "diagonal")
    if position not in {"diagonal", "center", "top", "bottom"}:
        position = "diagonal"

    # ── Save source file ───────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_watermarked.pdf")
    file.save(src_path)

    try:
        import fitz
        import math

        doc     = fitz.open(src_path)
        new_doc = fitz.open()
        new_doc.insert_pdf(doc)   # copy all pages
        doc.close()

        for i in range(new_doc.page_count):
            page   = new_doc.load_page(i)
            pw, ph = page.rect.width, page.rect.height

            # Font size: ~5 % of the shorter page dimension, clamped 18–96 pt
            font_size = max(18, min(96, round(min(pw, ph) * 0.05)))

            # Measure text width to centre it
            font = fitz.Font("helv")
            tw   = font.text_length(text, fontsize=font_size)

            if position == "diagonal":
                # Rotate 45° counter-clockwise around page centre
                cx = pw / 2
                cy = ph / 2
                angle_deg = 45.0
                angle_rad = math.radians(angle_deg)
                # Place baseline so text centre lands on page centre
                ox = cx - (tw / 2) * math.cos(angle_rad) + (font_size / 2) * math.sin(angle_rad)
                oy = cy + (tw / 2) * math.sin(angle_rad) + (font_size / 2) * math.cos(angle_rad)
                morph = (fitz.Point(cx, cy), fitz.Matrix(angle_deg))
            elif position == "center":
                ox = (pw - tw) / 2
                oy = ph / 2 + font_size / 2
                morph = None
            elif position == "top":
                ox = (pw - tw) / 2
                oy = ph * 0.15 + font_size
                morph = None
            else:   # bottom
                ox = (pw - tw) / 2
                oy = ph * 0.88
                morph = None

            page.insert_text(
                fitz.Point(ox, oy),
                text,
                fontname="helv",
                fontsize=font_size,
                color=(r, g, b),
                fill_opacity=opacity,
                stroke_opacity=opacity,
                morph=morph,
            )

        new_doc.save(out_path, garbage=4, deflate=True)
        new_doc.close()

        logger.info("Watermark PDF: text='%s', position=%s, uid=%s", text, position, uid)

    except Exception as exc:
        logger.error("Watermark PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except Exception:
                pass
        return jsonify({"error": "Failed to add watermark. The PDF may be encrypted or corrupted."}), 500

    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Watermark produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_watermarked.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception:
            pass
        return response

    return send_file(out_path, as_attachment=True, download_name=download_name)


# ── Rotate PDF route ───────────────────────────────────────────────────────

@app.route("/rotate-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def rotate_pdf_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Parameters ─────────────────────────────────────────────────────────
    # angle: 90 (CW), 180, 270 (CCW)
    try:
        angle = int(request.form.get("angle", "90"))
        if angle not in (90, 180, 270):
            angle = 90
    except ValueError:
        angle = 90

    # pages: "all" or a range string like "1-3, 5"
    pages_param = request.form.get("pages", "all").strip()

    # ── Save source ────────────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_rotated.pdf")
    file.save(src_path)

    pages_rotated = 0
    total_pages   = 0

    try:
        import fitz

        doc         = fitz.open(src_path)
        total_pages = doc.page_count

        if total_pages == 0:
            doc.close()
            return jsonify({"error": "PDF has no pages."}), 400

        # Determine which 0-based page indices to rotate
        if pages_param.lower() == "all" or pages_param == "":
            target_indices = set(range(total_pages))
        else:
            pairs = _parse_page_ranges(pages_param, total_pages)
            target_indices = set()
            for start, end in pairs:
                for i in range(start, end + 1):
                    target_indices.add(i)

        # Apply rotation: add angle to current rotation, keep in 0/90/180/270
        for i in target_indices:
            page         = doc.load_page(i)
            current      = page.rotation          # 0, 90, 180, or 270
            new_rotation = (current + angle) % 360
            page.set_rotation(new_rotation)
            pages_rotated += 1

        doc.save(out_path, garbage=4, deflate=True)
        doc.close()

        logger.info("Rotate PDF: %d/%d pages rotated %d°, uid=%s",
                    pages_rotated, total_pages, angle, uid)

    except ValueError as ve:
        logger.warning("Rotate PDF bad page range: %s", ve)
        return jsonify({"error": str(ve)}), 400

    except Exception as exc:
        logger.error("Rotate PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try: os.remove(out_path)
            except Exception: pass
        return jsonify({"error": "Failed to rotate PDF. The file may be encrypted or corrupted."}), 500

    finally:
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Rotation produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_rotated.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception: pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=download_name)
    response.headers["X-Pages-Rotated"] = str(pages_rotated)
    response.headers["X-Total-Pages"]   = str(total_pages)
    response.headers["X-Angle"]         = str(angle)
    response.headers["Access-Control-Expose-Headers"] = "X-Pages-Rotated, X-Total-Pages, X-Angle"
    return response


# ── Unlock PDF route ───────────────────────────────────────────────────────

@app.route("/unlock-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def unlock_pdf_route():
    # ── Validate file (lenient — broken header possible on encrypted PDFs) ─
    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"error": "No file provided."}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext != ".pdf":
        return jsonify({"error": "Please upload a PDF file (.pdf)."}), 400

    file.seek(0, 2); size = file.tell(); file.seek(0)
    if size == 0:
        return jsonify({"error": "File is empty."}), 400
    if size > app.config["MAX_FILE_SIZE"]:
        return jsonify({"error": f"File too large. Maximum {app.config['MAX_FILE_SIZE'] // (1024*1024)} MB."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Password ───────────────────────────────────────────────────────────
    password = request.form.get("password", "")   # may be empty for owner-lock only

    # ── Save source ────────────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_unlocked.pdf")
    file.save(src_path)

    try:
        import fitz

        doc = fitz.open(src_path)

        if doc.is_encrypted:
            # Try the provided password.  MuPDF tries both user and owner roles.
            auth_result = doc.authenticate(password)
            if auth_result == 0:
                doc.close()
                return jsonify({
                    "error": "Incorrect password. Please check and try again."
                }), 401
            logger.info("Unlock PDF: authenticated (result=%d) for %s", auth_result, safe_name)
        else:
            # PDF has no encryption — still save a clean copy (removes any
            # owner-level print/copy restrictions embedded without a password).
            logger.info("Unlock PDF: no encryption found, saving clean copy for %s", safe_name)

        # Save without any encryption — this is the "unlock"
        doc.save(
            out_path,
            garbage=4,
            deflate=True,
            encryption=fitz.PDF_ENCRYPT_NONE,   # strip all encryption
        )
        total_pages = doc.page_count
        doc.close()

    except Exception as exc:
        logger.error("Unlock PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try: os.remove(out_path)
            except Exception: pass
        return jsonify({"error": "Could not unlock this PDF. It may use unsupported encryption."}), 500

    finally:
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Unlock produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_unlocked.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception: pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=download_name)
    response.headers["X-Total-Pages"] = str(total_pages)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Pages"
    return response


# ── Protect PDF route ──────────────────────────────────────────────────────

@app.route("/protect-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def protect_pdf_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8); file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Password parameters ────────────────────────────────────────────────
    user_pw  = request.form.get("user_pw",  "").strip()
    owner_pw = request.form.get("owner_pw", "").strip()

    if not user_pw:
        return jsonify({"error": "A user (open) password is required."}), 400
    if len(user_pw) > 128 or len(owner_pw) > 128:
        return jsonify({"error": "Password too long (max 128 characters)."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # If owner password not set, make it the same as user password
    if not owner_pw:
        owner_pw = user_pw

    # ── Save source ────────────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_protected.pdf")
    file.save(src_path)

    try:
        import fitz

        # Allow printing; restrict modify/copy/annotations
        permissions = (
            fitz.PDF_PERM_PRINT
            | fitz.PDF_PERM_PRINT_HQ
            | fitz.PDF_PERM_ACCESSIBILITY
        )

        doc = fitz.open(src_path)

        # If source is already encrypted, try to open it without a password
        # (works for owner-locked-only PDFs); if it fails, ask user to unlock first
        if doc.is_encrypted:
            if doc.authenticate("") == 0:
                doc.close()
                return jsonify({
                    "error": "This PDF is already password-protected. "
                             "Please unlock it first, then re-protect it."
                }), 400

        total_pages = doc.page_count

        doc.save(
            out_path,
            encryption=fitz.PDF_ENCRYPT_AES_256,
            user_pw=user_pw,
            owner_pw=owner_pw,
            permissions=permissions,
            garbage=4,
            deflate=True,
        )
        doc.close()

        logger.info("Protect PDF: %d pages, uid=%s", total_pages, uid)

    except Exception as exc:
        logger.error("Protect PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try: os.remove(out_path)
            except Exception: pass
        return jsonify({"error": "Failed to protect the PDF. It may be corrupted or already encrypted."}), 500

    finally:
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Protection produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_protected.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception: pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=download_name)
    response.headers["X-Total-Pages"] = str(total_pages)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Pages"
    return response


# ── Page Numbers route ─────────────────────────────────────────────────────

@app.route("/page-numbers", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def page_numbers_route():
    # ── Validate file ──────────────────────────────────────────────────────
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8); file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    # ── Quota check ────────────────────────────────────────────────────────
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)

    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── Parameters ─────────────────────────────────────────────────────────
    # position: tl / tc / tr / bl / bc / br  (top/bottom · left/center/right)
    position = request.form.get("position", "bc")
    if position not in {"tl", "tc", "tr", "bl", "bc", "br"}:
        position = "bc"

    # format: "n" / "pn" / "nN" / "pnN"
    fmt = request.form.get("format", "n")
    if fmt not in {"n", "pn", "nN", "pnN"}:
        fmt = "n"

    try:
        start_num = max(1, int(request.form.get("start", "1")))
    except ValueError:
        start_num = 1

    # font_size: 8 – 24
    try:
        font_size = max(8, min(24, int(request.form.get("font_size", "11"))))
    except ValueError:
        font_size = 11

    # colour: hex string like "333333"
    raw_color = request.form.get("color", "333333").lstrip("#")
    try:
        r = int(raw_color[0:2], 16) / 255.0
        g = int(raw_color[2:4], 16) / 255.0
        b = int(raw_color[4:6], 16) / 255.0
    except Exception:
        r, g, b = 0.2, 0.2, 0.2

    # ── Save source ────────────────────────────────────────────────────────
    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_numbered.pdf")
    file.save(src_path)

    try:
        import fitz

        doc         = fitz.open(src_path)
        total_pages = doc.page_count

        if total_pages == 0:
            doc.close()
            return jsonify({"error": "PDF has no pages."}), 400

        margin = max(font_size * 1.4, 14.0)   # distance from edge in points

        for i in range(total_pages):
            page   = doc.load_page(i)
            pw, ph = page.rect.width, page.rect.height
            num    = start_num + i

            # Build label string
            if   fmt == "n":   label = str(num)
            elif fmt == "pn":  label = f"Page {num}"
            elif fmt == "nN":  label = f"{num} / {total_pages}"
            else:              label = f"Page {num} of {total_pages}"

            # Measure text width to centre horizontally
            font = fitz.Font("helv")
            tw   = font.text_length(label, fontsize=font_size)

            # Compute x, y based on position code
            row, col = position[0], position[1]   # t/b, l/c/r

            if   col == "l": x = margin
            elif col == "r": x = pw - tw - margin
            else:            x = (pw - tw) / 2      # centre

            # y is the text baseline
            if row == "t":
                y = margin + font_size
            else:
                y = ph - margin

            page.insert_text(
                fitz.Point(x, y),
                label,
                fontname="helv",
                fontsize=font_size,
                color=(r, g, b),
            )

        doc.save(out_path, garbage=4, deflate=True)
        doc.close()

        logger.info("Page numbers: %d pages, fmt=%s, pos=%s, uid=%s",
                    total_pages, fmt, position, uid)

    except Exception as exc:
        logger.error("Page numbers error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try: os.remove(out_path)
            except Exception: pass
        return jsonify({"error": "Failed to add page numbers. The PDF may be encrypted or corrupted."}), 500

    finally:
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Page numbering produced no output. Please try again."}), 500

    # ── Increment quota ────────────────────────────────────────────────────
    ConversionCounter.increment(fingerprint, request)

    base_name     = os.path.splitext(safe_name)[0]
    download_name = f"{base_name}_numbered.pdf"

    @after_this_request
    def remove_output(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception: pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=download_name)
    response.headers["X-Total-Pages"] = str(total_pages)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Pages"
    return response


# ── Translate PDF ──────────────────────────────────────────────────────────

TRANSLATE_TARGET_LANGS = [
    "English", "Arabic", "French", "Spanish", "German", "Italian",
    "Portuguese", "Russian", "Chinese (Simplified)", "Chinese (Traditional)",
    "Japanese", "Korean", "Turkish", "Dutch", "Polish", "Hindi",
    "Ukrainian", "Swedish", "Norwegian", "Danish",
    "Greek", "Hebrew", "Indonesian", "Vietnamese",
]

_TRANSLATE_FITZ_FONT = {
    "Chinese (Simplified)":  "china-s",
    "Chinese (Traditional)": "china-t",
    "Japanese": "japan",
    "Korean":   "korea",
}

_MYMEMORY_LANG_CODE = {
    "English": "en", "Arabic": "ar", "French": "fr", "Spanish": "es",
    "German": "de", "Italian": "it", "Portuguese": "pt", "Russian": "ru",
    "Chinese (Simplified)": "zh-CN", "Chinese (Traditional)": "zh-TW",
    "Japanese": "ja", "Korean": "ko", "Turkish": "tr", "Dutch": "nl",
    "Polish": "pl", "Hindi": "hi", "Ukrainian": "uk", "Swedish": "sv",
    "Norwegian": "no", "Danish": "da", "Greek": "el", "Hebrew": "he",
    "Indonesian": "id", "Vietnamese": "vi",
}
# Reverse: ISO code → full name (used to normalise frontend ISO codes back to full names)
_LANG_CODE_TO_NAME = {v: k for k, v in _MYMEMORY_LANG_CODE.items()}

_MYMEMORY_URL        = "https://api.mymemory.translated.net/get"
_MAX_CHUNK_SIZE      = 400
_CHUNK_DELAY         = 1.0   # seconds between chunks
_MAX_RETRIES         = 3


def _split_to_chunks(text, max_size=_MAX_CHUNK_SIZE):
    """Split text into chunks at sentence boundaries, each ≤ max_size chars."""
    # Normalise newlines into sentence-like splits
    sentences = []
    for line in text.replace('\n', ' \n ').split('\n'):
        for s in line.replace('. ', '.|').replace('! ', '!|').replace('? ', '?|').split('|'):
            if s.strip():
                sentences.append(s.strip())

    chunks, current = [], ""
    for sentence in sentences:
        if len(current) + len(sentence) + 1 <= max_size:
            current += (" " if current else "") + sentence
        else:
            if current:
                chunks.append(current)
            # If a single sentence exceeds max_size, hard-split it
            if len(sentence) > max_size:
                for i in range(0, len(sentence), max_size):
                    chunks.append(sentence[i:i + max_size])
                current = ""
            else:
                current = sentence
    if current:
        chunks.append(current)
    return chunks or [text[:max_size]]


def _translate_one_chunk(text, src_code, target_code):
    """Translate a single chunk via MyMemory with exponential-backoff retry."""
    for attempt in range(_MAX_RETRIES):
        try:
            resp = _requests.get(
                _MYMEMORY_URL,
                params={"q": text, "langpair": f"{src_code}|{target_code}"},
                timeout=15,
            )
            data = resp.json()
            translated = data.get("responseData", {}).get("translatedText", "")
            # Detect rate-limit signals
            if resp.status_code == 429 or data.get("responseStatus") == 429 \
                    or "MYMEMORY WARNING" in translated:
                wait = 3 * (attempt + 1)
                logger.warning("MyMemory rate limit hit, retrying in %ds (attempt %d)", wait, attempt + 1)
                time.sleep(wait)
                continue
            if data.get("responseStatus") == 200:
                return translated
            raise Exception(f"MyMemory error {data.get('responseStatus')}: {data.get('responseDetails', '')}")
        except _requests.RequestException:
            time.sleep(2)
    logger.warning("All retries failed for chunk, returning original text")
    return text   # graceful fallback — return untranslated chunk


def _mymemory_translate(text, target_code, src_code="en"):
    """Translate text via MyMemory, chunking at sentence boundaries."""
    if not text.strip():
        return text
    chunks = _split_to_chunks(text)
    out = []
    for idx, chunk in enumerate(chunks):
        if idx > 0:
            time.sleep(_CHUNK_DELAY)
        out.append(_translate_one_chunk(chunk, src_code, target_code))
    return " ".join(out)


@app.route("/extract-text-region", methods=["POST"])
@limiter.limit("30 per minute")
def extract_text_region():
    """Extract text from a rectangular region of a PDF page (percentage coordinates)."""
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    try:
        import json as _json
        regions = _json.loads(request.form.get("regions", "[]"))
    except Exception:
        return jsonify({"error": "Invalid regions format."}), 400

    if not regions:
        return jsonify({"error": "No regions provided."}), 400

    uid      = str(uuid.uuid4())
    src_path = os.path.join(app.config["UPLOAD_FOLDER"],
                            f"{uid}_{os.path.basename(file.filename)}")
    file.save(src_path)
    try:
        import fitz
        doc = fitz.open(src_path)
        parts = []
        for region in regions:
            page_idx = int(region.get("page", 0))
            if page_idx < 0 or page_idx >= len(doc):
                continue
            page = doc.load_page(page_idx)
            pw, ph = page.rect.width, page.rect.height
            x0 = float(region.get("x0_pct", 0))  / 100.0 * pw
            y0 = float(region.get("y0_pct", 0))  / 100.0 * ph
            x1 = float(region.get("x1_pct", 100)) / 100.0 * pw
            y1 = float(region.get("y1_pct", 100)) / 100.0 * ph
            rect = fitz.Rect(min(x0, x1), min(y0, y1), max(x0, x1), max(y0, y1))
            text = page.get_text("text", clip=rect).strip()
            if text:
                parts.append(text)
        doc.close()
    finally:
        try: os.remove(src_path)
        except: pass

    combined = "\n".join(parts)
    if not combined:
        return jsonify({"error": "No text found in the selected region. Try selecting a different area."}), 400

    return jsonify({"text": combined})


@app.route("/extract-pdf-paragraphs", methods=["POST"])
@limiter.limit("20 per minute")
def extract_pdf_paragraphs():
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    uid      = str(uuid.uuid4())
    src_path = os.path.join(app.config["UPLOAD_FOLDER"],
                            f"{uid}_{os.path.basename(file.filename)}")
    file.save(src_path)
    try:
        import fitz
        doc        = fitz.open(src_path)
        paragraphs = []
        for page_num in range(len(doc)):
            page   = doc.load_page(page_num)
            blocks = page.get_text("blocks")
            for block in blocks:
                if block[6] != 0:          # skip image blocks
                    continue
                text = block[4].strip()
                if not text or len(text) < 10:
                    continue
                for chunk in _split_to_chunks(text):
                    if chunk.strip():
                        paragraphs.append({"text": chunk.strip(), "page": page_num + 1})
        doc.close()
    finally:
        try: os.remove(src_path)
        except: pass

    if not paragraphs:
        return jsonify({"error": "No text found in PDF. This tool only works on text-based PDFs, not scanned images."}), 400

    return jsonify({"paragraphs": paragraphs, "total": len(paragraphs)})


@app.route("/translate", methods=["POST"])
@limiter.limit("60 per minute")
def translate_endpoint():
    """Translation via MyMemory free API.

    Request JSON body:
        { "text": "...", "from_code": "en", "to_code": "fr" }

    Response:
        { "translatedText": "..." }  on success
        { "error": "..." }           on failure
    """
    data      = request.get_json(silent=True) or {}
    text      = str(data.get("text", "")).strip()
    from_code = str(data.get("from_code", "en")).strip().lower()
    to_code   = str(data.get("to_code",   "en")).strip().lower()

    if not text:
        return jsonify({"error": "No text provided."}), 400
    if from_code == to_code:
        return jsonify({"translatedText": text})

    try:
        translated = _ts_translate_text(text, from_code, to_code)
        return jsonify({"translatedText": translated})
    except Exception as exc:
        logger.error("/translate error (%s→%s): %s", from_code, to_code, exc)
        return jsonify({"error": str(exc)}), 500


FREE_DAILY_TRANSLATIONS = 3

@app.route("/translate-chunk", methods=["POST"])
@limiter.limit("3 per minute; 10 per hour")  # Stricter limits
def translate_chunk():
    try:
        fingerprint = get_client_fingerprint(request)
        
        data        = request.get_json(silent=True) or {}
        text        = str(data.get("text", "")).strip()
        target_lang = str(data.get("target_lang", "English")).strip()
        from_lang   = str(data.get("from_lang", "en")).strip()

        if not text:
            return jsonify({"error": "No text provided."}), 400

        # Normalize both language names and shorthand codes
        normalized_lang = _LANG_CODE_TO_NAME.get(target_lang) or _LANG_CODE_TO_NAME.get(target_lang.lower())
        target_lang = normalized_lang or target_lang

        # Resolve from_lang ISO code → full name → MyMemory code
        normalized_from = _LANG_CODE_TO_NAME.get(from_lang) or _LANG_CODE_TO_NAME.get(from_lang.lower())
        from_name = normalized_from or from_lang
        from_code = _MYMEMORY_LANG_CODE.get(from_name) or _MYMEMORY_LANG_CODE.get(from_name.lower()) or from_lang.lower()

        # ✅ Check JWT Pro token
        pro_token = request.cookies.get("pro_token")
        is_pro = False
        if pro_token:
            valid, payload = ProToken.verify(pro_token)
            if valid and payload.get("fingerprint") == fingerprint:
                is_pro = True

        if not is_pro:
            # ✅ Check server-side daily quota
            allowed, used, remaining = TranslationQuota.check_and_increment(
                fingerprint, 
                limit=FREE_DAILY_TRANSLATIONS
            )

            if not allowed:
                return jsonify({
                    "error":   "free_limit_reached",
                    "used":    used,
                    "limit":   FREE_DAILY_TRANSLATIONS,
                    "message": f"You've used your {FREE_DAILY_TRANSLATIONS} free daily translations.",
                }), 402

        target_code = _MYMEMORY_LANG_CODE.get(target_lang) or _MYMEMORY_LANG_CODE.get(target_lang.lower()) or target_lang.lower()
        translated  = _ts_translate_text(text, from_code, target_code)

        remaining = None
        if not is_pro:
            remaining = FREE_DAILY_TRANSLATIONS - used

        is_rtl = target_lang in _RTL_LANG_NAMES or target_code in {"ar", "he", "ar-SA", "he-IL", "ar-sa", "he-il"}
        return jsonify({"translatedText": translated, "remaining": remaining, "isRtl": is_rtl})

    except Exception as exc:
        logger.error("translate-chunk error: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/translate-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def translate_pdf_route():
    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)
    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8); file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    target_lang = request.form.get("target_lang", "English").strip()
    # Normalise: if frontend sent an ISO code (e.g. "fr"), convert to full name ("French")
    target_lang = _LANG_CODE_TO_NAME.get(target_lang, target_lang)
    if target_lang not in TRANSLATE_TARGET_LANGS:
        target_lang = "English"
    target_code = _MYMEMORY_LANG_CODE.get(target_lang, "en")

    from_lang_raw = request.form.get("from_lang", "en").strip()
    from_lang_name = _LANG_CODE_TO_NAME.get(from_lang_raw, from_lang_raw)
    from_code = _MYMEMORY_LANG_CODE.get(from_lang_name, from_lang_raw.lower())

    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_translated.pdf")
    file.save(src_path)

    total_pages = 0
    try:
        import fitz

        doc = fitz.open(src_path)
        total_pages = doc.page_count

        if total_pages == 0:
            doc.close()
            return jsonify({"error": "PDF has no pages."}), 400

        MAX_PAGES = 25
        if total_pages > MAX_PAGES:
            doc.close()
            return jsonify({"error": f"PDF has {total_pages} pages. Maximum is {MAX_PAGES} for translation."}), 400

        # Optional page selection from frontend dialog
        trans_pages_param = request.form.get("trans_pages", "").strip()
        if trans_pages_param:
            try:
                pairs = _parse_page_ranges(trans_pages_param, total_pages)
                selected_indices = sorted({i for s, e in pairs for i in range(s, e + 1)})
            except Exception:
                selected_indices = list(range(total_pages))
        else:
            selected_indices = list(range(total_pages))

        pages_text = []
        for i in range(total_pages):
            if i in selected_indices:
                pages_text.append(doc.load_page(i).get_text().strip())
            else:
                pages_text.append(None)  # skip this page
        doc.close()

        if not any(t for t in pages_text if t is not None):
            return jsonify({"error": "No text found in PDF. This tool only works on text-based PDFs, not scanned images."}), 400

        fitz_font  = _TRANSLATE_FITZ_FONT.get(target_lang, "helv")
        rtl        = _is_rtl_lang(target_lang)

        # Path to bundled Arabic font (used for all RTL languages)
        _amiri_path = os.path.join(app.root_path, "static", "fonts", "Amiri-Regular.ttf")
        use_amiri   = rtl and os.path.exists(_amiri_path) and _RTL_LIBS_OK

        # Fallback system font search (only for non-Amiri RTL or CJK)
        font_file = None
        if not use_amiri:
            _candidates = [
                "C:/Windows/Fonts/arialuni.ttf",
                "C:/Windows/Fonts/tahoma.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf",
                "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            ]
            if target_lang in {"Arabic", "Hebrew"} or fitz_font != "helv":
                for _p in _candidates:
                    if os.path.exists(_p):
                        font_file = _p
                        break

        translated_pages = []
        for page_text in pages_text:
            if page_text is None:
                translated_pages.append(None)  # keep original page (not selected)
                continue
            if not page_text.strip():
                translated_pages.append("")
                continue
            translated_pages.append(_ts_translate_text(page_text, from_code, target_code))

        # ── Build output PDF ─────────────────────────────────────────────────
        import textwrap as _textwrap

        new_doc = fitz.open()
        A4_W, A4_H = 595.0, 842.0
        MARGIN     = 50.0
        FONTSIZE   = 11
        LINE_H     = FONTSIZE * 1.45
        USABLE_W   = A4_W - 2 * MARGIN
        USABLE_H   = A4_H - 2 * MARGIN

        CHARS_PER_LINE = max(40, int(USABLE_W / (FONTSIZE * 0.55)))
        LINES_PER_PAGE = max(10, int(USABLE_H / LINE_H))

        def _paginate_ltr(text):
            lines = []
            for para in text.split("\n"):
                if para.strip():
                    lines.extend(_textwrap.wrap(para, width=CHARS_PER_LINE) or [""])
                lines.append("")
            chunks = []
            for start in range(0, max(1, len(lines)), LINES_PER_PAGE):
                chunks.append("\n".join(lines[start:start + LINES_PER_PAGE]))
            return chunks or [""]

        def _paginate_rtl(text):
            """Split RTL text into page-sized character blocks (no textwrap)."""
            MAX_CHARS = 1800
            if len(text) <= MAX_CHARS:
                return [text]
            chunks, current = [], ""
            for line in text.split("\n"):
                line = line.strip()
                if not line:
                    continue
                if len(current) + len(line) + 1 <= MAX_CHARS:
                    current += ("\n" if current else "") + line
                else:
                    if current:
                        chunks.append(current)
                    current = line
            if current:
                chunks.append(current)
            return chunks or [text]

        def _add_page(text_chunk):
            pg   = new_doc.new_page(width=A4_W, height=A4_H)
            rect = fitz.Rect(MARGIN, MARGIN, A4_W - MARGIN, A4_H - MARGIN)
            if use_amiri:
                # RTL: reshape + reorder, then right-align with Amiri font
                display = _prepare_rtl_text(text_chunk)
                pg.insert_font(fontname="Amiri", fontfile=_amiri_path)
                rv = pg.insert_textbox(rect, display, fontname="Amiri",
                                       fontsize=FONTSIZE, color=(0.05, 0.05, 0.05), align=2)
                if rv < 0:
                    pg.insert_textbox(rect, display, fontname="Amiri",
                                      fontsize=9, color=(0.05, 0.05, 0.05), align=2)
            else:
                kwargs = dict(fontsize=FONTSIZE, color=(0.05, 0.05, 0.05), align=0)
                try:
                    if font_file:
                        kwargs["fontfile"] = font_file
                        kwargs["fontname"] = "custom"
                    else:
                        kwargs["fontname"] = fitz_font
                    rv = pg.insert_textbox(rect, text_chunk, **kwargs)
                    if rv < 0:
                        logger.warning("Translate: insert_textbox overflow rv=%.1f", rv)
                except Exception as _fe:
                    logger.warning("Translate font error, fallback helv: %s", _fe)
                    try:
                        pg.insert_textbox(rect, text_chunk, fontname="helv",
                                          fontsize=FONTSIZE, color=(0.05, 0.05, 0.05), align=0)
                    except Exception:
                        pass

        _paginate = _paginate_rtl if rtl else _paginate_ltr

        for t_text in translated_pages:
            if t_text is None:
                continue  # page was not selected for translation — skip
            if not t_text:
                _add_page("")
                continue
            for chunk in _paginate(t_text):
                _add_page(chunk)

        new_doc.save(out_path, garbage=4, deflate=True)
        new_doc.close()
        logger.info("Translate PDF: %d src pages, lang=%s, uid=%s",
                    total_pages, str(target_lang), str(uid))

    except Exception as exc:
        logger.error("Translate PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try: os.remove(out_path)
            except Exception: pass
        err_str = str(exc)
        return jsonify({"error": f"Translation failed: {err_str[:200]}"}), 500
    finally:
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Translation produced no output."}), 500

    ConversionCounter.increment(fingerprint, request)

    base_name  = os.path.splitext(safe_name)[0]
    lang_slug  = target_lang.lower().replace(" ", "_").replace("(", "").replace(")", "")
    dl_name    = f"{base_name}_translated_{lang_slug}.pdf"

    @after_this_request
    def _rm_translate(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception: pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=dl_name)
    response.headers["X-Total-Pages"] = str(total_pages)
    response.headers["X-Target-Lang"] = target_lang
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Pages, X-Target-Lang"
    return response


# ── Sign PDF ───────────────────────────────────────────────────────────────

@app.route("/sign-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def sign_pdf_route():
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8); file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    sig_data_url = request.form.get("signature_data", "").strip()
    sig_file     = request.files.get("signature_file")
    if not sig_data_url and (not sig_file or sig_file.filename == ""):
        return jsonify({"error": "No signature provided. Draw or upload a signature first."}), 400

    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)
    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    position = request.form.get("position", "br")
    if position not in {"tl", "tc", "tr", "bl", "bc", "br"}:
        position = "br"

    # Click-based placement (x_pct / y_pct from frontend page-click modal)
    try:
        sig_x_pct = float(request.form.get("x_pct", ""))
        sig_y_pct = float(request.form.get("y_pct", ""))
        use_click_pos = 0.0 <= sig_x_pct <= 100.0 and 0.0 <= sig_y_pct <= 100.0
    except (ValueError, TypeError):
        sig_x_pct = sig_y_pct = 0.0
        use_click_pos = False

    pages_opt    = request.form.get("pages", "last")
    pages_custom = request.form.get("pages_custom", "")

    SIZE_MAP = {"small": 0.18, "medium": 0.28, "large": 0.38}
    size_factor = SIZE_MAP.get(request.form.get("size", "medium"), 0.28)

    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_signed.pdf")
    file.save(src_path)

    signed_count = 0
    total_pages  = 0
    try:
        import fitz, base64 as _base64, io
        from PIL import Image as PILImage

        if sig_data_url:
            if "," not in sig_data_url:
                return jsonify({"error": "Invalid signature data format."}), 400
            try:
                sig_bytes = _base64.b64decode(sig_data_url.split(",", 1)[1])
            except Exception:
                return jsonify({"error": "Failed to decode signature image."}), 400
        else:
            sig_bytes = sig_file.read()

        img = PILImage.open(io.BytesIO(sig_bytes)).convert("RGBA")
        sig_w_px, sig_h_px = img.size
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        sig_png = buf.getvalue()

        doc = fitz.open(src_path)
        total_pages = doc.page_count

        if pages_opt == "first":
            target_indices = [0]
        elif pages_opt == "last":
            target_indices = [total_pages - 1]
        elif pages_opt == "all":
            target_indices = list(range(total_pages))
        else:
            try:
                pairs = _parse_page_ranges(pages_custom, total_pages)
                target_indices = sorted({i for s, e in pairs for i in range(s, e + 1)})
            except Exception:
                target_indices = [total_pages - 1]

        if use_click_pos:
            try:
                click_page = max(0, min(int(request.form.get("page_num", "0")), total_pages - 1))
            except (ValueError, TypeError):
                click_page = total_pages - 1
            target_indices = [click_page]

        MARGIN = 20.0
        for i in target_indices:
            page   = doc.load_page(i)
            pw, ph = page.rect.width, page.rect.height
            sig_w  = pw * size_factor
            sig_h  = sig_w * (sig_h_px / max(sig_w_px, 1))

            if use_click_pos:
                x = (sig_x_pct / 100.0) * pw - sig_w / 2
                y = (sig_y_pct / 100.0) * ph - sig_h / 2
                x = max(0.0, min(x, pw - sig_w))
                y = max(0.0, min(y, ph - sig_h))
            else:
                row, col = position[0], position[1]
                if   col == "l": x = MARGIN
                elif col == "r": x = pw - sig_w - MARGIN
                else:            x = (pw - sig_w) / 2
                if row == "t":   y = MARGIN
                else:            y = ph - sig_h - MARGIN

            page.insert_image(fitz.Rect(x, y, x + sig_w, y + sig_h), stream=sig_png)
            signed_count += 1

        doc.save(out_path, garbage=4, deflate=True)
        doc.close()
        logger.info("Sign PDF: %d/%d pages signed, pos=%s, uid=%s",
                    signed_count, total_pages, position, uid)

    except Exception as exc:
        logger.error("Sign PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try: os.remove(out_path)
            except Exception: pass
        return jsonify({"error": "Failed to sign PDF."}), 500
    finally:
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Signing produced no output."}), 500

    ConversionCounter.increment(fingerprint, request)

    base_name = os.path.splitext(safe_name)[0]

    @after_this_request
    def _rm_sign(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception: pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=f"{base_name}_signed.pdf")
    response.headers["X-Pages-Signed"] = str(signed_count)
    response.headers["X-Total-Pages"]  = str(total_pages)
    response.headers["Access-Control-Expose-Headers"] = "X-Pages-Signed, X-Total-Pages"
    return response


# ── PDF Page Preview (for edit-pdf visual placement) ──────────────────────

@app.route("/pdf-page-preview", methods=["POST"])
@limiter.limit("30 per minute")
def pdf_page_preview():
    """Render a single PDF page to JPEG and return it as base64.
    Does NOT consume a quota slot — only used for the edit-pdf UI preview."""
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    try:
        page_num = max(0, int(request.form.get("page", "0")))
    except (ValueError, TypeError):
        page_num = 0

    uid      = str(uuid.uuid4())
    src_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_editpreview.pdf")
    file.save(src_path)

    try:
        import fitz, base64 as _b64

        doc         = fitz.open(src_path)
        total_pages = doc.page_count

        if total_pages == 0:
            doc.close()
            return jsonify({"error": "PDF has no pages."}), 400

        page_num = min(page_num, total_pages - 1)
        page     = doc.load_page(page_num)
        page_w   = page.rect.width
        page_h   = page.rect.height

        mat = fitz.Matrix(1.5, 1.5)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
        try:
            img_bytes = pix.tobytes("jpeg", jpg_quality=80)
        except Exception:
            img_bytes = pix.tobytes("png")
        doc.close()

    except Exception as exc:
        logger.error("PDF page preview error: %s", exc, exc_info=True)
        return jsonify({"error": "Failed to render page preview."}), 500
    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    return jsonify({
        "image":       _b64.b64encode(img_bytes).decode(),
        "page":        page_num,
        "total_pages": total_pages,
        "page_width":  page_w,
        "page_height": page_h,
    })


# ── PDF Text Extract (for edit-pdf span overlay) ──────────────────────────

@app.route("/pdf-text-extract", methods=["POST"])
@limiter.limit("30 per minute")
def pdf_text_extract():
    """Extract text spans with bounding boxes from a single PDF page.
    Returns coordinates as percentages of page dimensions.
    Does NOT consume a quota slot."""
    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8)
    file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    try:
        page_num = max(0, int(request.form.get("page", "0")))
    except (ValueError, TypeError):
        page_num = 0

    uid      = str(uuid.uuid4())
    src_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_textextract.pdf")
    file.save(src_path)

    try:
        import fitz

        doc         = fitz.open(src_path)
        total_pages = doc.page_count
        if total_pages == 0:
            doc.close()
            return jsonify({"error": "PDF has no pages."}), 400

        page_num = min(page_num, total_pages - 1)
        page     = doc.load_page(page_num)
        page_w   = page.rect.width
        page_h   = page.rect.height

        raw   = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)
        spans = []
        for block in raw.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    text = span.get("text", "").strip()
                    if not text:
                        continue
                    bx0, by0, bx1, by1 = span["bbox"]
                    if (bx1 - bx0) < 1 or (by1 - by0) < 1:
                        continue
                    spans.append({
                        "text":      text,
                        "x0_pct":    round(bx0 / page_w * 100, 3),
                        "y0_pct":    round(by0 / page_h * 100, 3),
                        "x1_pct":    round(bx1 / page_w * 100, 3),
                        "y1_pct":    round(by1 / page_h * 100, 3),
                        "font_size": round(span.get("size", 12), 1),
                        "color_hex": "%06x" % (span.get("color", 0) & 0xFFFFFF),
                    })
                    if len(spans) >= 500:
                        break
                if len(spans) >= 500:
                    break
            if len(spans) >= 500:
                break
        doc.close()

    except Exception as exc:
        logger.error("PDF text extract error: %s", exc, exc_info=True)
        return jsonify({"error": "Failed to extract text."}), 500
    finally:
        if os.path.exists(src_path):
            try:
                os.remove(src_path)
            except Exception:
                pass

    return jsonify({
        "page":       page_num,
        "page_width": page_w,
        "page_height": page_h,
        "spans":      spans,
    })


# ── RTL/Arabic text helper ────────────────────────────────────────────────

def _prepare_text_for_pdf(text):
    """Reshape and apply BiDi algorithm to Arabic/RTL text.
    Returns (display_text, is_rtl)."""
    import re
    if re.search(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]', text):
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(text)), True
        except Exception:
            pass
    return text, False


# ── Edit PDF ───────────────────────────────────────────────────────────────

@app.route("/edit-pdf", methods=["POST"])
@limiter.limit(os.getenv("CONVERT_RATE_LIMIT", "10 per minute"))
def edit_pdf_route():
    import json as _json
    from collections import defaultdict

    file = request.files.get("file")
    is_valid, error_msg = validate_file(file, [".pdf"], app.config["MAX_FILE_SIZE"])
    if not is_valid:
        return jsonify({"error": error_msg}), 400

    header = file.read(8); file.seek(0)
    if not header.startswith(b"%PDF"):
        return jsonify({"error": "File does not appear to be a valid PDF."}), 400

    fingerprint = get_client_fingerprint(request)
    _cc_used, _cc_budget, _, _cc_pro = ConversionCounter.get_status(fingerprint)
    if _cc_used >= _cc_budget and not _cc_pro:
        return jsonify({
            "error":              "quota_exceeded",
            "message":            "You've used all your free conversions. Upgrade to continue.",
            "conversions_used":   _cc_used,
            "conversions_budget": _cc_budget,
        }), 402

    # ── New multi-change path ──────────────────────────────────────────────
    changes_raw = request.form.get("changes", "")
    if changes_raw:
        try:
            changes = _json.loads(changes_raw)
            if not isinstance(changes, list) or len(changes) == 0:
                return jsonify({"error": "No changes provided."}), 400
            if len(changes) > 50:
                return jsonify({"error": "Too many changes (max 50)."}), 400
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid changes data."}), 400

        # Validate each change
        valid_actions = {"add", "replace", "delete"}
        for ch in changes:
            if ch.get("action") not in valid_actions:
                return jsonify({"error": "Invalid action in changes."}), 400
            for key in ("x0_pct", "y0_pct", "x1_pct", "y1_pct"):
                if ch.get("action") in ("replace", "delete") and key in ch:
                    v = float(ch[key])
                    if not (0.0 <= v <= 100.0):
                        return jsonify({"error": f"Out-of-range value for {key}."}), 400
            for text_key in ("text", "new_text"):
                if text_key in ch and len(ch[text_key]) > 500:
                    return jsonify({"error": "Text too long (max 500 characters)."}), 400

        uid       = str(uuid.uuid4())
        safe_name = os.path.basename(file.filename)
        src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
        out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_edited.pdf")
        file.save(src_path)

        edited_count = 0
        total_pages  = 0
        try:
            import fitz

            doc = fitz.open(src_path)
            total_pages = doc.page_count
            if total_pages == 0:
                doc.close()
                return jsonify({"error": "PDF has no pages."}), 400

            # Group changes by page
            page_changes = defaultdict(list)
            for ch in changes:
                pg = max(0, min(int(ch.get("page", 0)), total_pages - 1))
                page_changes[pg].append(ch)

            helv_font = fitz.Font("helv")

            for page_idx, ch_list in page_changes.items():
                page   = doc.load_page(page_idx)
                pw, ph = page.rect.width, page.rect.height

                # Step 1: add redact annotations for delete/replace
                needs_redact = False
                for ch in ch_list:
                    if ch["action"] in ("delete", "replace"):
                        x0 = float(ch["x0_pct"]) / 100 * pw
                        y0 = float(ch["y0_pct"]) / 100 * ph
                        x1 = float(ch["x1_pct"]) / 100 * pw
                        y1 = float(ch["y1_pct"]) / 100 * ph
                        page.add_redact_annot(fitz.Rect(x0, y0, x1, y1), fill=(1, 1, 1))
                        needs_redact = True

                # Step 2: apply redactions once per page
                if needs_redact:
                    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)

                # Step 3: insert new text for add/replace
                # Group by color to use one TextWriter per color
                color_groups = defaultdict(list)
                for ch in ch_list:
                    if ch["action"] in ("add", "replace"):
                        color_groups[ch.get("color", "000000")].append(ch)

                for color_hex, chs in color_groups.items():
                    raw_color = color_hex.lstrip("#")
                    try:
                        rc = int(raw_color[0:2], 16) / 255.0
                        gc = int(raw_color[2:4], 16) / 255.0
                        bc = int(raw_color[4:6], 16) / 255.0
                    except Exception:
                        rc = gc = bc = 0.0
                    tw = fitz.TextWriter(page.rect)
                    for ch in chs:
                        font_size = max(8, min(72, int(ch.get("font_size", 12))))
                        new_text = ch.get("text") or ch.get("new_text", "")
                        if not new_text:
                            continue
                        display_text, is_rtl = _prepare_text_for_pdf(new_text)
                        text_w = helv_font.text_length(display_text, fontsize=font_size)
                        if ch["action"] == "add":
                            lx = float(ch["x_pct"]) / 100 * pw
                            ly = float(ch["y_pct"]) / 100 * ph + font_size
                            if is_rtl:
                                lx = lx - text_w  # right-align from click point
                        else:  # replace
                            ly = float(ch["y1_pct"]) / 100 * ph
                            if is_rtl:
                                lx = float(ch["x1_pct"]) / 100 * pw - text_w
                            else:
                                lx = float(ch["x0_pct"]) / 100 * pw
                        tw.append(fitz.Point(lx, ly), display_text,
                                  font=helv_font, fontsize=font_size)
                    tw.write_text(page, color=(rc, gc, bc))

                edited_count += 1

            doc.save(out_path, garbage=4, deflate=True)
            doc.close()
            logger.info("Edit PDF (multi-change): %d pages, %d changes, uid=%s",
                        edited_count, len(changes), uid)

        except Exception as exc:
            logger.error("Edit PDF error: %s", exc, exc_info=True)
            if os.path.exists(out_path):
                try: os.remove(out_path)
                except Exception: pass
            return jsonify({"error": "Failed to edit PDF."}), 500
        finally:
            if os.path.exists(src_path):
                try: os.remove(src_path)
                except Exception: pass

        if not os.path.exists(out_path):
            return jsonify({"error": "Edit produced no output."}), 500

        ConversionCounter.increment(fingerprint, request)
        base_name = os.path.splitext(safe_name)[0]

        @after_this_request
        def _rm_edit_multi(response):
            try:
                if os.path.exists(out_path): os.remove(out_path)
            except Exception: pass
            return response

        resp = send_file(out_path, as_attachment=True, download_name=f"{base_name}_edited.pdf")
        resp.headers["X-Pages-Edited"] = str(edited_count)
        resp.headers["X-Total-Pages"]  = str(total_pages)
        resp.headers["Access-Control-Expose-Headers"] = "X-Pages-Edited, X-Total-Pages"
        return resp

    # ── Legacy single-text path (fallback) ────────────────────────────────
    text = request.form.get("text", "").strip()
    if not text:
        return jsonify({"error": "No text provided."}), 400
    if len(text) > 500:
        return jsonify({"error": "Text is too long (max 500 characters)."}), 400

    position = request.form.get("position", "tl")
    if position not in {"tl", "tc", "tr", "bl", "bc", "br"}:
        position = "tl"

    try:
        x_pct = float(request.form.get("x_pct", ""))
        y_pct = float(request.form.get("y_pct", ""))
        use_coords = 0.0 <= x_pct <= 100.0 and 0.0 <= y_pct <= 100.0
    except (ValueError, TypeError):
        x_pct = y_pct = 0.0
        use_coords = False

    pages_str = request.form.get("pages", "all").strip()
    try:
        font_size = max(8, min(72, int(request.form.get("font_size", "12"))))
    except ValueError:
        font_size = 12

    raw_color = request.form.get("color", "000000").lstrip("#")
    try:
        rc = int(raw_color[0:2], 16) / 255.0
        gc = int(raw_color[2:4], 16) / 255.0
        bc = int(raw_color[4:6], 16) / 255.0
    except Exception:
        rc, gc, bc = 0.0, 0.0, 0.0

    uid       = str(uuid.uuid4())
    safe_name = os.path.basename(file.filename)
    src_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{safe_name}")
    out_path  = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_edited.pdf")
    file.save(src_path)

    edited_count = 0
    total_pages  = 0
    try:
        import fitz

        doc = fitz.open(src_path)
        total_pages = doc.page_count
        if total_pages == 0:
            doc.close()
            return jsonify({"error": "PDF has no pages."}), 400

        if pages_str.lower() in ("all", "", "*"):
            target_indices = list(range(total_pages))
        else:
            try:
                pairs = _parse_page_ranges(pages_str, total_pages)
                target_indices = sorted({i for s, e in pairs for i in range(s, e + 1)})
            except Exception:
                target_indices = list(range(total_pages))

        if use_coords:
            try:
                click_page = max(0, min(int(request.form.get("page_num", "0")), total_pages - 1))
            except (ValueError, TypeError):
                click_page = 0
            target_indices = [click_page]

        margin = max(font_size * 1.4, 14.0)
        lines  = text.split("\n")
        line_h = font_size * 1.4
        row, col = position[0], position[1]
        helv_font = fitz.Font("helv")
        line_widths = [helv_font.text_length(ln, fontsize=font_size) if ln else 0.0
                       for ln in lines]

        for i in target_indices:
            page   = doc.load_page(i)
            pw, ph = page.rect.width, page.rect.height
            tw = fitz.TextWriter(page.rect)

            for j, line in enumerate(lines):
                if not line:
                    continue
                display_line, is_rtl = _prepare_text_for_pdf(line)
                line_w = helv_font.text_length(display_line, fontsize=font_size)
                if use_coords:
                    lx = (x_pct / 100.0) * pw
                    ly = (y_pct / 100.0) * ph + font_size + j * line_h
                    if is_rtl:
                        lx = lx - line_w
                else:
                    if   col == "l": lx = margin
                    elif col == "r": lx = pw - line_w - margin
                    else:            lx = (pw - line_w) / 2
                    if row == "t":
                        ly = margin + font_size + j * line_h
                    else:
                        ly = ph - margin - (len(lines) - 1 - j) * line_h
                tw.append(fitz.Point(lx, ly), display_line, font=helv_font, fontsize=font_size)

            tw.write_text(page, color=(rc, gc, bc))
            edited_count += 1

        doc.save(out_path, garbage=4, deflate=True)
        doc.close()
        logger.info("Edit PDF: %d pages edited, coords=%s, pos=%s, uid=%s",
                    edited_count, use_coords, position, uid)

    except Exception as exc:
        logger.error("Edit PDF error: %s", exc, exc_info=True)
        if os.path.exists(out_path):
            try: os.remove(out_path)
            except Exception: pass
        return jsonify({"error": "Failed to edit PDF."}), 500
    finally:
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    if not os.path.exists(out_path):
        return jsonify({"error": "Edit produced no output."}), 500

    ConversionCounter.increment(fingerprint, request)
    base_name = os.path.splitext(safe_name)[0]

    @after_this_request
    def _rm_edit(response):
        try:
            if os.path.exists(out_path): os.remove(out_path)
        except Exception: pass
        return response

    response = send_file(out_path, as_attachment=True, download_name=f"{base_name}_edited.pdf")
    response.headers["X-Pages-Edited"] = str(edited_count)
    response.headers["X-Total-Pages"]  = str(total_pages)
    response.headers["Access-Control-Expose-Headers"] = "X-Pages-Edited, X-Total-Pages"
    return response


# ── Apply translation voucher (unlocks pro_unlocked flag) ──────────────────

@app.route("/apply-voucher", methods=["POST"])
@limiter.limit("10 per minute")
def apply_voucher():
    """Validate a voucher code and unlock pro translation for this session."""
    data = request.get_json(silent=True) or {}
    code = str(data.get("code", "")).strip().upper()
    if not code:
        return jsonify({"success": False, "message": "Please enter a voucher code."}), 400
    valid_codes = _load_voucher_codes()
    if not valid_codes:
        return jsonify({"success": False, "message": "Voucher system is not enabled on this server."}), 503
    if code not in valid_codes:
        return jsonify({"success": False, "message": "Invalid voucher code. Please check and try again."}), 400
    redeemed = session.get("redeemed_vouchers", [])
    if code in redeemed:
        return jsonify({"success": False, "message": "This voucher has already been used in this session."}), 400
    session["pro_unlocked"]      = True
    session["redeemed_vouchers"] = redeemed + [code]
    session.modified = True
    logger.info("Translation voucher applied: code=%s", code)
    return jsonify({"success": True, "message": "Pro unlocked — enjoy unlimited translations!"})


# ── Redeem voucher ─────────────────────────────────────────────────────────

@app.route("/redeem-voucher", methods=["POST"])
@limiter.limit("5 per minute; 10 per hour")  # Stricter limits
def redeem_voucher():
    fingerprint = get_client_fingerprint(request)
    
    # ✅ Check attempt rate limit and lockout
    allowed, message = VoucherSecurity.check_attempt(fingerprint)
    if not allowed:
        return jsonify({"error": message, "locked": True}), 429  # Too Many Requests
    
    data = request.get_json(silent=True) or {}
    code = str(data.get("code", "")).strip().upper()

    if not code:
        return jsonify({"error": "Please enter a voucher code."}), 400

    valid_codes = _load_voucher_codes()
    if not valid_codes:
        return jsonify({"error": "Voucher system is not enabled on this server."}), 503

    if code not in valid_codes:
        VoucherSecurity.record_attempt(fingerprint, False)  # ✅ Record failed attempt
        return jsonify({"error": "Invalid voucher code. Please check and try again."}), 400

    # Prevent double-redeem in session
    redeemed = session.get("redeemed_vouchers", [])
    if code in redeemed:
        VoucherSecurity.record_attempt(fingerprint, False)
        return jsonify({"error": "This voucher has already been redeemed in this session."}), 400

    # Grant conversions server-side via ConversionCounter
    ConversionCounter.grant_pro(fingerprint, VOUCHER_GRANT)
    _v_used, new_budget, _v_rem, _ = ConversionCounter.get_status(fingerprint)
    remaining = new_budget - _v_used

    session["redeemed_vouchers"] = redeemed + [code]
    session.modified = True

    VoucherSecurity.record_attempt(fingerprint, True)  # ✅ Record success

    logger.info("Voucher redeemed: code=%s, granted=%d", LogSanitizer.sanitize(code), VOUCHER_GRANT)

    return jsonify({
        "success":    True,
        "granted":    VOUCHER_GRANT,
        "remaining":  remaining,
        "budget":     new_budget,
    })


# ── Error handlers ─────────────────────────────────────────────────────────

@app.errorhandler(429)
def ratelimit_error(e):
    logger.warning("Rate limit exceeded from %s", get_remote_address())
    return jsonify({"error": "Too many requests. Please wait a moment and try again."}), 429


@app.errorhandler(413)
def file_too_large(e):
    return jsonify({"error": f"File too large. Maximum size is {app.config['MAX_FILE_SIZE'] // (1024*1024)} MB."}), 413


# ── SEO Routes ────────────────────────────────────────────────────────────

@app.route('/sitemap.xml')
def sitemap():
    return send_from_directory('static', 'sitemap.xml', mimetype='application/xml')


@app.route('/robots.txt')
def robots():
    return """User-agent: *
Allow: /
Sitemap: https://convertly-web.onrender.com/sitemap.xml
""", 200, {'Content-Type': 'text/plain'}


@app.route("/privacy")
def privacy():
    """Privacy Policy page."""
    return render_template("privacy.html")


@app.route("/support", methods=["GET", "POST"])
def support():
    """Support / feedback page."""
    if request.method == "GET":
        return render_template("support.html", sent=False, error=None)

    subject_type = request.form.get("subject_type", "").strip()
    message      = request.form.get("message", "").strip()
    user_email   = request.form.get("user_email", "").strip()

    if not subject_type or not message:
        return render_template("support.html", sent=False,
                               error="Please fill in all required fields.")

    # ── Save message to SQLite ─────────────────────────────────────────────
    import sqlite3 as _sqlite3, tempfile as _tempfile

    _db_path = os.getenv("QUOTA_DB_PATH", os.path.join(_tempfile.gettempdir(), "quota.db"))
    try:
        with _sqlite3.connect(_db_path, timeout=10) as _conn:
            _conn.execute("""
                CREATE TABLE IF NOT EXISTS support_messages (
                    id        INTEGER PRIMARY KEY AUTOINCREMENT,
                    created   TEXT    NOT NULL,
                    topic     TEXT    NOT NULL,
                    email     TEXT,
                    message   TEXT    NOT NULL
                )
            """)
            _conn.execute(
                "INSERT INTO support_messages (created, topic, email, message) VALUES (?,?,?,?)",
                (_datetime.utcnow().isoformat(sep=" ", timespec="seconds"),
                 subject_type, user_email or None, message)
            )
        app.logger.info("Support message saved to DB.")
    except Exception as exc:
        app.logger.error("Could not save support message: %s", exc)

    # ── Send email via Gmail API (HTTPS, works on Render) ─────────────────
    def _send_gmail():
        try:
            import urllib.request, urllib.parse, json as _json, base64 as _b64
            from email.mime.text import MIMEText

            client_id     = os.getenv("GMAIL_CLIENT_ID", "")
            client_secret = os.getenv("GMAIL_CLIENT_SECRET", "")
            refresh_token = os.getenv("GMAIL_REFRESH_TOKEN", "")
            sender        = os.getenv("GMAIL_SENDER", "")
            to_addr       = os.getenv("SUPPORT_EMAIL", "ibrahimezzeldinmirghani@gmail.com")

            if not all([client_id, client_secret, refresh_token, sender]):
                app.logger.warning("Support email: Gmail API env vars not set")
                return

            # 1. Get a fresh access token
            token_data = urllib.parse.urlencode({
                "client_id":     client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
                "grant_type":    "refresh_token",
            }).encode()
            try:
                with urllib.request.urlopen(
                    urllib.request.Request("https://oauth2.googleapis.com/token",
                                           data=token_data, method="POST"),
                    timeout=15
                ) as r:
                    access_token = _json.loads(r.read())["access_token"]
                    app.logger.info("Gmail token obtained OK")
            except urllib.error.HTTPError as e:
                app.logger.error("Gmail token refresh failed %s: %s", e.code, e.read().decode())
                return

            # 2. Build and send the email
            body = f"Topic: {subject_type}\n"
            if user_email:
                body += f"Reply-to: {user_email}\n"
            body += f"\n{message}"

            msg = MIMEText(body, "plain")
            msg["From"]    = sender
            msg["To"]      = to_addr
            msg["Subject"] = f"[Convertly Support] {subject_type}"
            if user_email:
                msg["Reply-To"] = user_email

            raw = _b64.urlsafe_b64encode(msg.as_bytes()).decode()
            try:
                with urllib.request.urlopen(
                    urllib.request.Request(
                        "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
                        data=_json.dumps({"raw": raw}).encode(),
                        headers={"Authorization": f"Bearer {access_token}",
                                 "Content-Type": "application/json"},
                        method="POST",
                    ),
                    timeout=15
                ) as r:
                    app.logger.info("Support email sent OK via Gmail API")
            except urllib.error.HTTPError as e:
                app.logger.error("Gmail API send failed %s: %s", e.code, e.read().decode())

        except Exception as exc:
            app.logger.error("Gmail API send failed: %s", exc)

    threading.Thread(target=_send_gmail, daemon=True).start()

    return render_template("support.html", sent=True, error=None)


# ── Admin: view support messages ──────────────────────────────────────────

@app.route("/admin/messages")
def admin_messages():
    import sqlite3 as _sqlite3, tempfile as _tempfile, hmac as _hmac, base64 as _b64
    admin_user = os.getenv("ADMIN_USERNAME", "admin")
    admin_pw   = os.getenv("ADMIN_PASSWORD", "")

    # Require Basic Auth — credentials in header, never in URL
    auth = request.headers.get("Authorization", "")
    authed = False
    if admin_pw and auth.startswith("Basic "):
        try:
            decoded = _b64.b64decode(auth[6:]).decode("utf-8")
            req_user, req_pass = decoded.split(":", 1)
            authed = (
                _hmac.compare_digest(req_user, admin_user) and
                _hmac.compare_digest(req_pass, admin_pw)
            )
        except Exception:
            pass
    if not authed:
        return ("Unauthorized", 401, {"WWW-Authenticate": 'Basic realm="Convertly Admin"'})

    _db_path = os.getenv("QUOTA_DB_PATH", os.path.join(_tempfile.gettempdir(), "quota.db"))
    rows = []
    try:
        with _sqlite3.connect(_db_path, timeout=10) as _conn:
            _conn.row_factory = _sqlite3.Row
            rows = _conn.execute(
                "SELECT id, created, topic, email, message FROM support_messages ORDER BY id DESC"
            ).fetchall()
    except Exception:
        pass

    html = ["<html><head><meta charset='utf-8'><title>Support Messages</title>",
            "<style>body{font-family:sans-serif;padding:24px;max-width:860px;margin:auto}",
            "table{width:100%;border-collapse:collapse}",
            "th,td{text-align:left;padding:8px 12px;border-bottom:1px solid #ddd;vertical-align:top}",
            "th{background:#f5f5f5}pre{white-space:pre-wrap;margin:0}</style></head><body>",
            f"<h2>Support Messages ({len(rows)})</h2>",
            "<table><tr><th>#</th><th>Date</th><th>Topic</th><th>Email</th><th>Message</th></tr>"]
    for r in rows:
        html.append(f"<tr><td>{r['id']}</td><td>{r['created']}</td><td>{r['topic']}</td>"
                    f"<td>{r['email'] or '—'}</td><td><pre>{r['message']}</pre></td></tr>")
    html.append("</table></body></html>")
    return "\n".join(html)


# ── Cache headers for static files ────────────────────────────────────────

@app.after_request
def add_cache_headers(response):
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000'
    return response


if __name__ == "__main__":
    debug_mode = os.getenv("FLASK_DEBUG", "False").lower() == "true"
    port       = int(os.getenv("FLASK_PORT", 5000))
    host       = os.getenv("HOST", "0.0.0.0")
    app.run(debug=debug_mode, host=host, port=port)
